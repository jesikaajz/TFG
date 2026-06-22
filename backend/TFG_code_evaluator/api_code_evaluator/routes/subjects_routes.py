from fastapi import APIRouter, Depends, HTTPException,  Query,  File, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from .teacher_assignment_routes import TeacherAssignment
from typing import Optional
from .course_offerings_routes import CourseOffering
from ..bd.connection import get_db
from ..models.subjects_model import Subject
from ..schemas.subject_schema import SubjectCreate, SubjectPatch, SubjectResponse, BulkSubjectResult
from ..dependencies.auth_dependencies import admin_required, get_current_user
from ..models.academic_year_model import AcademicYear
from .exercises_routes import Exercise
from ..models.user_model import User
from ..models.enrollment_model import Enrollment
from ..models.enrollment_detail_model import EnrollmentDetail
from openpyxl import load_workbook, Workbook
from io import BytesIO
from sqlalchemy import func, distinct, case
from ..models.submissions_model import Submission
from ..models.evaluation_model import Evaluation
from ..models.exercise_model import Exercise
from ..models.enrollment_model import Enrollment
from ..models.enrollment_detail_model import EnrollmentDetail

def teacher_has_access(db: Session, user_id: int, offering_id: int) -> bool:
    """
    Verifica si un profesor tiene acceso a una oferta de curso específica.
    """
    assignment = db.query(TeacherAssignment).filter(
        TeacherAssignment.professor_id == user_id,
        TeacherAssignment.course_offering_id == offering_id
    ).first()
    
    return assignment is not None

router = APIRouter(prefix="/subjects", tags=["Subjects"])

@router.get("/", response_model=list[SubjectResponse])
def get_all(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(Subject).all()


@router.post("/", response_model=SubjectResponse, status_code=status.HTTP_201_CREATED)
def create(
    subject: SubjectCreate,
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    # 1. Recortar el nombre y validar que no esté vacío
    name_trimmed = subject.name.strip() if subject.name else ""
    if not name_trimmed:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El nombre de la asignatura no puede estar vacío"
        )

    # 2. Verificar duplicados (insensible a mayúsculas y espacios)
    existing = db.query(Subject).filter(
        func.lower(func.trim(Subject.name)) == func.lower(name_trimmed)
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Ya existe una asignatura con el nombre '{name_trimmed}'"
        )

    # 3. Crear la asignatura con el nombre recortado
    new_subject = Subject(name=name_trimmed, description=subject.description)
    db.add(new_subject)
    db.commit()
    db.refresh(new_subject)
    return new_subject

@router.patch("/{subject_id}", response_model=SubjectResponse)
def patch(
    subject_id: int,
    data: SubjectPatch,
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    subject = db.query(Subject).filter(Subject.id == subject_id).first()
    if not subject:
        raise HTTPException(404, "Subject not found")

    update_data = data.dict(exclude_unset=True)

    # Validar nombre si viene en la actualización
    if "name" in update_data:
        name_trimmed = update_data["name"].strip() if update_data["name"] else ""
        if not name_trimmed:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El nombre de la asignatura no puede estar vacío"
            )
        # Verificar duplicados, excluyendo el sujeto actual
        existing = db.query(Subject).filter(
            Subject.id != subject_id,
            func.lower(func.trim(Subject.name)) == func.lower(name_trimmed)
        ).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Ya existe otra asignatura con el nombre '{name_trimmed}'"
            )
        update_data["name"] = name_trimmed  # Guardamos recortado

    # Aplicar actualizaciones
    for k, v in update_data.items():
        setattr(subject, k, v)

    db.commit()
    db.refresh(subject)
    return subject


@router.delete("/{subject_id}")
def delete(subject_id: int, db: Session = Depends(get_db), _: User = Depends(admin_required)):
    subject = db.query(Subject).filter(Subject.id == subject_id).first()
    if not subject:
        raise HTTPException(404, "Subject not found")

    db.delete(subject)
    db.commit()
    return {"message": "Subject deleted"}

@router.get("/exercises/subject/{subject_id}")
def get_exercises_by_subject(
    subject_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener todos los ejercicios de una asignatura/curso específico.
    Busca todas las ofertas de la asignatura y luego los ejercicios de cada oferta.
    """
    # Verificar que la asignatura existe
    subject = db.query(Subject).filter(Subject.id == subject_id).first()
    if not subject:
        raise HTTPException(404, f"Subject with id {subject_id} not found")
    
    # Obtener todas las ofertas de esta asignatura
    offerings = db.query(CourseOffering).filter(
        CourseOffering.subject_id == subject_id
    ).all()
    
    offering_ids = [offering.id for offering in offerings]
    
    if not offering_ids:
        return []  # No hay ofertas, no hay ejercicios
    
    # Construir query base
    query = db.query(Exercise).filter(
        Exercise.course_offering_id.in_(offering_ids)
    )
    
    # Filtrar según rol
    if current_user.role == "student":
        query = query.filter(Exercise.visibility == True)
    elif current_user.role == "teacher":
        # Verificar que el profesor tiene acceso a al menos una oferta
        has_access = False
        for offering_id in offering_ids:
            if teacher_has_access(db, current_user.id, offering_id):
                has_access = True
                break
        if not has_access:
            raise HTTPException(403, "Not authorized to view exercises for this subject")
    # admin puede ver todo
    
    exercises = query.all()
    
    # Construir respuesta
    result = []
    for exercise in exercises:
        # Obtener la oferta para información adicional
        offering = db.query(CourseOffering).filter(
            CourseOffering.id == exercise.course_offering_id
        ).first()
        
        academic_year = None
        if offering:
            year = db.query(AcademicYear).filter(
                AcademicYear.id == offering.academic_year_id
            ).first()
            if year:
                academic_year = f"{year.start_year}-{year.end_year}"
        
        result.append({
            "id": exercise.id,
            "title": exercise.title,
            "description": exercise.description,
            "deadline": exercise.deadline,
            "course_offering_id": exercise.course_offering_id,
            "visibility": exercise.visibility,
            "academic_year": academic_year,
            "has_solution": exercise.solution is not None,
            "test_cases_count": len(exercise.test_cases) if exercise.test_cases else 0
        })
    
    return result

@router.get("/subjects/count")
async def get_count_all_subjects(
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    """Obtener el número total de materias - Solo administradores"""
    # Contar materias directamente
    total_subjects = db.query(Subject).count()
    
    return {"count": total_subjects}

# ---------------------------------------------------------
# Obtener ofertas de asignatura asignadas a un profesor
# ---------------------------------------------------------
@router.get("/course-offerings/by-teacher")
def get_course_offerings_by_teacher(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    teacher_id: Optional[int] = Query(
        None,
        description="ID del profesor (solo administradores)"
    )
):
    """
    Obtener las ofertas de asignatura impartidas por un profesor.

    - Los administradores pueden consultar las ofertas de cualquier profesor.
    - Los profesores solo pueden consultar sus propias ofertas.
    - Los estudiantes pueden consultar las ofertas de cualquier profesor (para ver qué cursos imparte cada uno).
    """

    # Verificar permisos
    if current_user.role not in ["teacher", "admin", "student"]:
        raise HTTPException(
            status_code=403,
            detail="No autorizado"
        )

    # Determinar profesor objetivo
    if current_user.role == "admin" and teacher_id is not None:
        target_teacher_id = teacher_id
    elif current_user.role == "student":
        # Los estudiantes pueden consultar cualquier profesor
        if teacher_id is None:
            raise HTTPException(
                status_code=400,
                detail="Los estudiantes deben especificar teacher_id"
            )
        target_teacher_id = teacher_id
    else:
        # Profesor solo puede ver sus propias ofertas
        target_teacher_id = current_user.id

    # Verificar que el usuario exista y sea profesor
    teacher = db.query(User).filter(
        User.id == target_teacher_id,
        User.role == "teacher"
    ).first()

    if not teacher:
        raise HTTPException(
            status_code=404,
            detail="Profesor no encontrado"
        )

    # Consultar ofertas asignadas
    offerings = db.query(
        CourseOffering.id.label("offering_id"),
        Subject.id.label("subject_id"),
        Subject.name.label("subject_name"),
        AcademicYear.start_year,
        AcademicYear.end_year
    ).join(
        TeacherAssignment,
        CourseOffering.id == TeacherAssignment.course_offering_id
    ).join(
        Subject,
        CourseOffering.subject_id == Subject.id
    ).join(
        AcademicYear,
        CourseOffering.academic_year_id == AcademicYear.id
    ).filter(
        TeacherAssignment.professor_id == target_teacher_id
    ).all()

    return [
        {
            "offering_id": offering.offering_id,
            "subject_id": offering.subject_id,
            "subject_name": offering.subject_name,
            "academic_year": f"{offering.start_year}-{offering.end_year}"
        }
        for offering in offerings
    ]

# ---------------------------------------------------------
# Obtener estudiantes matriculados en una oferta de asignatura
@router.get("/course-offerings/{offering_id}/students")
def get_students_by_course_offering(
    offering_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener estudiantes matriculados en una oferta concreta.

    - Los administradores pueden consultar cualquier oferta.
    - Los profesores solo pueden consultar ofertas que imparten.
    - Los estudiantes solo pueden consultar ofertas en las que están matriculados.
    """

    # Verificar existencia de oferta
    offering = db.query(CourseOffering).filter(
        CourseOffering.id == offering_id
    ).first()

    if not offering:
        raise HTTPException(
            status_code=404,
            detail="Oferta de asignatura no encontrada"
        )

    # Verificar permisos según rol
    if current_user.role == "admin":
        # Admin puede ver cualquier oferta
        pass
        
    elif current_user.role == "teacher":
        # Profesor solo puede ver ofertas que imparte
        assignment = db.query(TeacherAssignment).filter(
            TeacherAssignment.professor_id == current_user.id,
            TeacherAssignment.course_offering_id == offering_id
        ).first()

        if not assignment:
            raise HTTPException(
                status_code=403,
                detail="No tienes acceso a esta oferta de asignatura"
            )
            
    elif current_user.role == "student":
        # Estudiante solo puede ver ofertas en las que está matriculado
        is_enrolled = db.query(EnrollmentDetail).join(
            Enrollment, EnrollmentDetail.enrollment_id == Enrollment.id
        ).filter(
            EnrollmentDetail.offering_id == offering_id,
            Enrollment.student_id == current_user.id
        ).first()
        
        if not is_enrolled:
            raise HTTPException(
                status_code=403,
                detail="No estás matriculado en esta oferta de asignatura"
            )
    else:
        raise HTTPException(
            status_code=403,
            detail="No autorizado"
        )

    # Obtener estudiantes matriculados (todos los estudiantes, no solo el usuario actual)
    students = db.query(
        User.id,
        User.name,
        User.email
    ).join(
        Enrollment, User.id == Enrollment.student_id
    ).join(
        EnrollmentDetail, Enrollment.id == EnrollmentDetail.enrollment_id
    ).filter(
        EnrollmentDetail.offering_id == offering_id,
        User.role == "student"
    ).distinct().all()

    return [
        {
            "id": student.id,
            "name": student.name,
            "email": student.email
        }
        for student in students
    ]
# ==========================================
# BULK CREATE SUBJECTS FROM EXCEL
# ==========================================

def _parse_excel_subjects(file: UploadFile) -> list[tuple[int, dict]]:
    allowed_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload a .xlsx file."
        )

    content = file.file.read()
    max_size = 5 * 1024 * 1024
    if len(content) > max_size:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="Excel file too large. Maximum size is 5 MB."
        )

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unable to read Excel file. Verify the file is a valid .xlsx workbook."
        )

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or not any(rows[0]):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file must have a header row with name and description."
        )

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    required_fields = {"name"}
    valid_fields = {"name", "description"}
    missing = required_fields - set(headers)
    unknown = set(headers) - valid_fields

    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required columns: {', '.join(sorted(missing))}."
        )
    if unknown:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unknown columns found: {', '.join(sorted(unknown))}."
        )

    parsed = []
    for row_index, row in enumerate(rows[1:], start=2):
        if not any(cell is not None for cell in row):
            continue

        row_data = {}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            row_data[header] = row[col_index] if col_index < len(row) else None

        parsed.append((row_index, row_data))

    return parsed


@router.post("/upload", response_model=BulkSubjectResult, status_code=status.HTTP_201_CREATED)
def upload_subjects(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    """
    Upload an Excel file (.xlsx) to create subjects in bulk.
    Columns must include: name, and optionally description.
    """
    rows = _parse_excel_subjects(file)
    created_subjects = []
    errors = []

    for row_number, row_data in rows:
        row_data = {
            key: (str(value).strip() if value is not None else None)
            for key, value in row_data.items()
        }

        name = row_data.get("name")
        description = row_data.get("description")

        if not name:
            errors.append({
                "row": row_number,
                "status": "error",
                "detail": "Subject name cannot be empty"
            })
            continue

        # Check if subject already exists
        existing = db.query(Subject).filter(Subject.name == name).first()
        if existing:
            errors.append({
                "row": row_number,
                "status": "error",
                "detail": f"Subject '{name}' already exists"
            })
            continue

        try:
            new_subject = Subject(
                name=name,
                description=description
            )
            db.add(new_subject)
            db.commit()
            db.refresh(new_subject)
            created_subjects.append(new_subject)
        except Exception as exc:
            db.rollback()
            errors.append({
                "row": row_number,
                "status": "error",
                "detail": str(exc)
            })

    return BulkSubjectResult(
        created_count=len(created_subjects),
        created=created_subjects,
        errors=errors
    )


@router.get("/template", status_code=200)
def download_subjects_template(_: User = Depends(admin_required)):
    """
    Download an Excel template for bulk subject upload.
    Columns: name (required), description (optional).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "plantilla_materias"
    headers = ["name", "description"]
    ws.append(headers)
    ws.append(["Algorithms", "Introduction to algorithms and data structures"])
    ws.append(["Data Structures", "Advanced data structures and their applications"])
    ws.append(["Programming Fundamentals", "Basic programming concepts and logic"])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = "plantilla_excel_subjects.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



# ==========================================
# RANKING DE ESTUDIANTES POR ASIGNATURA
# ==========================================
@router.get("/{subject_id}/ranking")
def get_subject_ranking(
    subject_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0)
):
    """
    Obtener ranking de estudiantes por asignatura.
    """
    
    try:
        # 1. Verificar que la asignatura existe
        subject = db.query(Subject).filter(Subject.id == subject_id).first()
        if not subject:
            raise HTTPException(404, f"Subject with id {subject_id} not found")
        
        # 2. Verificar permisos
        if current_user.role not in ["admin", "teacher"]:
            if current_user.role == "student":
                # Verificar si el estudiante está matriculado
                is_enrolled = db.query(EnrollmentDetail).join(
                    CourseOffering, EnrollmentDetail.offering_id == CourseOffering.id
                ).filter(
                    CourseOffering.subject_id == subject_id,
                    EnrollmentDetail.enrollment.has(Enrollment.student_id == current_user.id)
                ).first()
                
                if not is_enrolled:
                    raise HTTPException(403, "No tienes acceso al ranking de esta asignatura")
            else:
                raise HTTPException(403, "Not authorized")
        
        # 3. Obtener todas las course offerings de la asignatura
        offerings = db.query(CourseOffering).filter(
            CourseOffering.subject_id == subject_id
        ).all()
        
        offering_ids = [o.id for o in offerings]
        
        if not offering_ids:
            return {
                "subject_id": subject_id,
                "subject_name": subject.name,
                "total_students": 0,
                "total_exercises": 0,
                "ranking": [],
                "current_user_rank": None,
                "pagination": {"limit": limit, "offset": offset, "has_more": False}
            }
        
        # 4. Obtener todos los ejercicios de esas offerings
        exercises = db.query(Exercise).filter(
            Exercise.course_offering_id.in_(offering_ids)
        ).all()
        
        exercise_ids = [e.id for e in exercises]
        total_exercises = len(exercise_ids)
        
        if not exercise_ids:
            return {
                "subject_id": subject_id,
                "subject_name": subject.name,
                "total_students": 0,
                "total_exercises": 0,
                "ranking": [],
                "current_user_rank": None,
                "pagination": {"limit": limit, "offset": offset, "has_more": False}
            }
        
        # 5. Obtener estudiantes matriculados (usando SQL directo para evitar problemas)
        # Esta es una versión más simple usando SQLAlchemy puro
        
        # Primero, obtener los IDs de los estudiantes matriculados
        student_ids_query = db.query(Enrollment.student_id).join(
            EnrollmentDetail, Enrollment.id == EnrollmentDetail.enrollment_id
        ).join(
            CourseOffering, EnrollmentDetail.offering_id == CourseOffering.id
        ).filter(
            CourseOffering.subject_id == subject_id
        ).distinct()
        
        student_ids = [s[0] for s in student_ids_query.all()]
        
        if not student_ids:
            return {
                "subject_id": subject_id,
                "subject_name": subject.name,
                "total_students": 0,
                "total_exercises": total_exercises,
                "ranking": [],
                "current_user_rank": None,
                "pagination": {"limit": limit, "offset": offset, "has_more": False}
            }
        
        # 6. Calcular ranking para cada estudiante
        ranking_data = []
        
        for student_id in student_ids:
            # Obtener información del estudiante
            student = db.query(User).filter(User.id == student_id).first()
            if not student:
                continue
            
            # Obtener todas las submissions del estudiante para estos ejercicios
            submissions = db.query(Submission).filter(
                Submission.student_id == student_id,
                Submission.exercise_id.in_(exercise_ids)
            ).all()
            
            # Calcular mejores notas por ejercicio
            best_scores = {}
            for sub in submissions:
                evaluation = db.query(Evaluation).filter(
                    Evaluation.submission_id == sub.id
                ).first()
                
                if evaluation and evaluation.score is not None:
                    if sub.exercise_id not in best_scores or evaluation.score > best_scores[sub.exercise_id]:
                        best_scores[sub.exercise_id] = evaluation.score
            
            # Contar ejercicios aprobados (score >= 60)
            solved = 0
            total_score = 0
            for score in best_scores.values():
                if score >= 60:
                    solved += 1
                total_score += score
            
            avg_score = total_score / len(best_scores) if best_scores else 0
            
            ranking_data.append({
                "student_id": student_id,
                "name": student.name,
                "email": student.email,
                "solved_exercises": solved,
                "avg_score": avg_score
            })
        
        # 7. Ordenar ranking
        ranking_data.sort(key=lambda x: (-x["solved_exercises"], -x["avg_score"]))
        
        # 8. Aplicar paginación
        total_students = len(ranking_data)
        paginated_ranking = ranking_data[offset:offset + limit]
        
        # 9. Construir respuesta con ranks
        ranking = []
        for idx, student in enumerate(paginated_ranking, start=offset + 1):
            ranking.append({
                "rank": idx,
                "student_id": student["student_id"],
                "name": student["name"],
                "email": student["email"],
                "solved_exercises": student["solved_exercises"],
                "total_exercises": total_exercises,
                "progress_percentage": round((student["solved_exercises"] / total_exercises * 100), 2) if total_exercises > 0 else 0,
                "average_score": round(student["avg_score"], 2)
            })
        
        # 10. Posición del usuario actual
        current_user_rank = None
        if current_user.role == "student":
            for idx, student in enumerate(ranking_data, start=1):
                if student["student_id"] == current_user.id:
                    current_user_rank = idx
                    break
        
        return {
            "subject_id": subject_id,
            "subject_name": subject.name,
            "total_students": total_students,
            "total_exercises": total_exercises,
            "ranking": ranking,
            "current_user_rank": current_user_rank,
            "pagination": {
                "limit": limit,
                "offset": offset,
                "has_more": offset + limit < total_students
            }
        }
        
    except Exception as e:
        print(f"Error in get_subject_ranking: {str(e)}")
        raise HTTPException(500, f"Internal server error: {str(e)}")

# ==========================================
# RANKING DEL ESTUDIANTE EN UNA ASIGNATURA ESPECÍFICA
# ==========================================
@router.get("/{subject_id}/my-ranking")
def get_my_subject_ranking(
    subject_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener la posición y estadísticas del estudiante actual en una asignatura.
    """
    
    try:
        if current_user.role != "student":
            raise HTTPException(403, "Solo estudiantes pueden ver su propio ranking")
        
        # Verificar que el estudiante está matriculado
        is_enrolled = db.query(EnrollmentDetail).join(
            CourseOffering, EnrollmentDetail.offering_id == CourseOffering.id
        ).filter(
            CourseOffering.subject_id == subject_id,
            EnrollmentDetail.enrollment.has(Enrollment.student_id == current_user.id)
        ).first()
        
        if not is_enrolled:
            raise HTTPException(403, "No estás matriculado en esta asignatura")
        
        # Obtener todas las course offerings
        offerings = db.query(CourseOffering).filter(
            CourseOffering.subject_id == subject_id
        ).all()
        
        offering_ids = [o.id for o in offerings]
        
        if not offering_ids:
            return {
                "subject_id": subject_id,
                "message": "No hay ejercicios en esta asignatura aún"
            }
        
        # Obtener ejercicios
        exercises = db.query(Exercise).filter(
            Exercise.course_offering_id.in_(offering_ids)
        ).all()
        
        exercise_ids = [e.id for e in exercises]
        total_exercises = len(exercise_ids)
        
        if not exercise_ids:
            return {
                "subject_id": subject_id,
                "message": "No hay ejercicios en esta asignatura aún"
            }
        
        # Obtener mejores notas del estudiante
        best_scores = {}
        submissions = db.query(Submission).filter(
            Submission.student_id == current_user.id,
            Submission.exercise_id.in_(exercise_ids)
        ).all()
        
        for sub in submissions:
            evaluation = db.query(Evaluation).filter(
                Evaluation.submission_id == sub.id
            ).first()
            
            if evaluation and evaluation.score is not None:
                if sub.exercise_id not in best_scores or evaluation.score > best_scores[sub.exercise_id]:
                    best_scores[sub.exercise_id] = evaluation.score
        
        # Calcular estadísticas
        solved_exercises = sum(1 for score in best_scores.values() if score >= 60)
        avg_score = sum(best_scores.values()) / len(best_scores) if best_scores else 0
        
        # Obtener todos los estudiantes matriculados
        student_ids = db.query(Enrollment.student_id).join(
            EnrollmentDetail, Enrollment.id == EnrollmentDetail.enrollment_id
        ).join(
            CourseOffering, EnrollmentDetail.offering_id == CourseOffering.id
        ).filter(
            CourseOffering.subject_id == subject_id
        ).distinct().all()
        
        student_ids = [s[0] for s in student_ids]
        
        # Calcular posición del estudiante
        position = 1
        if len(student_ids) > 1:
            student_scores = []
            for student_id in student_ids:
                if student_id == current_user.id:
                    continue
                
                # Contar ejercicios aprobados de otros estudiantes
                submissions_other = db.query(Submission).filter(
                    Submission.student_id == student_id,
                    Submission.exercise_id.in_(exercise_ids)
                ).all()
                
                other_scores = {}
                for sub in submissions_other:
                    evaluation = db.query(Evaluation).filter(
                        Evaluation.submission_id == sub.id
                    ).first()
                    
                    if evaluation and evaluation.score is not None and evaluation.score >= 60:
                        if sub.exercise_id not in other_scores or evaluation.score > other_scores[sub.exercise_id]:
                            other_scores[sub.exercise_id] = evaluation.score
                
                solved_other = len(other_scores)
                student_scores.append((student_id, solved_other))
            
            # Ordenar y encontrar posición
            student_scores.sort(key=lambda x: -x[1])
            for idx, (sid, _) in enumerate(student_scores, start=1):
                if solved_exercises < _:
                    position = idx + 1
                else:
                    position = idx
                    break
        
        subject_name = db.query(Subject.name).filter(Subject.id == subject_id).scalar()
        
        return {
            "subject_id": subject_id,
            "subject_name": subject_name,
            "student_id": current_user.id,
            "student_name": current_user.name,
            "solved_exercises": solved_exercises,
            "total_exercises": total_exercises,
            "progress_percentage": round((solved_exercises / total_exercises * 100), 2) if total_exercises > 0 else 0,
            "average_score": round(avg_score, 2),
            "rank": position,
            "total_students": len(student_ids),
            "exercises_detail": [
                {
                    "exercise_id": ex_id,
                    "best_score": score,
                    "passed": score >= 60
                }
                for ex_id, score in best_scores.items()
            ]
        }
        
    except Exception as e:
        print(f"Error in get_my_subject_ranking: {str(e)}")
        raise HTTPException(500, f"Internal server error: {str(e)}")

# ==========================================
# RANKING DE UN PROFESOR (ver ranking de sus estudiantes)
# ==========================================
@router.get("/teacher/my-students-ranking")
def get_teacher_students_ranking(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0)
):
    """
    Obtener ranking de TODOS los estudiantes del profesor.
    Agrupa por estudiante, sumando ejercicios aprobados en todas las asignaturas que imparte.
    """
    
    if current_user.role not in ["teacher", "admin"]:
        raise HTTPException(403, "Solo profesores o administradores pueden ver este ranking")
    
    # Obtener todas las course offerings del profesor (si es teacher, solo las suyas)
    if current_user.role == "teacher":
        offerings = db.query(CourseOffering).join(
            TeacherAssignment, CourseOffering.id == TeacherAssignment.course_offering_id
        ).filter(
            TeacherAssignment.professor_id == current_user.id
        ).all()
    else:
        offerings = db.query(CourseOffering).all()
    
    offering_ids = [o.id for o in offerings]
    
    if not offering_ids:
        return {
            "total_students": 0,
            "ranking": []
        }
    
    # Obtener ejercicios de esas offerings
    exercises = db.query(Exercise).filter(
        Exercise.course_offering_id.in_(offering_ids)
    ).all()
    
    exercise_ids = [e.id for e in exercises]
    
    if not exercise_ids:
        return {
            "total_students": 0,
            "ranking": []
        }
    
    # Obtener estudiantes matriculados en esas offerings
    enrolled_students = db.query(User.id, User.name, User.email).join(
        Enrollment, User.id == Enrollment.student_id
    ).join(
        EnrollmentDetail, Enrollment.id == EnrollmentDetail.enrollment_id
    ).filter(
        EnrollmentDetail.offering_id.in_(offering_ids),
        User.role == "student"
    ).distinct().all()
    
    student_ids = [s.id for s in enrolled_students]
    
    if not student_ids:
        return {
            "total_students": 0,
            "ranking": []
        }
    
    # Subconsulta para mejores notas
    best_scores_subquery = db.query(
        Submission.student_id,
        Submission.exercise_id,
        func.max(Evaluation.score).label('best_score')
    ).join(
        Evaluation, Submission.id == Evaluation.submission_id
    ).filter(
        Submission.exercise_id.in_(exercise_ids),
        Submission.student_id.in_(student_ids),
        Evaluation.score >= 60
    ).group_by(
        Submission.student_id,
        Submission.exercise_id
    ).subquery()
    
    # Ranking
    ranking_query = db.query(
        User.id,
        User.name,
        User.email,
        func.count(distinct(best_scores_subquery.c.exercise_id)).label('solved_exercises'),
        func.avg(best_scores_subquery.c.best_score).label('avg_score')
    ).join(
        best_scores_subquery, User.id == best_scores_subquery.c.student_id
    ).filter(
        User.id.in_(student_ids)
    ).group_by(
        User.id, User.name, User.email
    ).order_by(
        func.count(distinct(best_scores_subquery.c.exercise_id)).desc(),
        func.avg(best_scores_subquery.c.best_score).desc()
    ).offset(offset).limit(limit).all()
    
    total_students = len(student_ids)
    total_exercises = len(exercise_ids)
    
    ranking = []
    for i, student in enumerate(ranking_query, start=offset + 1):
        ranking.append({
            "rank": i,
            "student_id": student.id,
            "name": student.name,
            "email": student.email,
            "solved_exercises": student.solved_exercises,
            "total_exercises": total_exercises,
            "progress_percentage": round((student.solved_exercises / total_exercises * 100), 2) if total_exercises > 0 else 0,
            "average_score": round(float(student.avg_score), 2) if student.avg_score else 0
        })
    
    return {
        "teacher_id": current_user.id,
        "teacher_name": current_user.name,
        "total_students": total_students,
        "total_exercises": total_exercises,
        "ranking": ranking,
        "pagination": {
            "limit": limit,
            "offset": offset,
            "has_more": offset + limit < total_students
        }
    }