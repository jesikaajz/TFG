from io import BytesIO

from fastapi import HTTPException
from fastapi import APIRouter, Depends, File, UploadFile, status
from sqlalchemy.orm import Session
from openpyxl import load_workbook, Workbook
from fastapi.responses import StreamingResponse

from ..schemas.user_schema import (
    UserCreate,
    UserPatch,
    UserResponse,
    UserRole,
    UserPublicRegister,
    PasswordChange,
    PasswordChangeWithCurrent,  # NUEVO IMPORT
    UserBulkCreate,
    BulkUploadResult,
)
from ..security.jwt_handler import hash_password, verify_password  # Agregar verify_password
from ..services.email_service import generate_random_password, send_temporary_password
from ..bd.connection import get_db
from ..models.user_model import User
from ..dependencies.auth_dependencies import admin_required
from .auth_routes import get_current_user
from ..models.enrollment_detail_model import EnrollmentDetail
from ..models.enrollment_model import Enrollment
from ..models.teacher_assignments_model import TeacherAssignment

router = APIRouter()

# ==========================================
# NUEVO ENDPOINT: REGISTRO DE USUARIO (PÚBLICO)
# ==========================================
@router.post("/register", response_model=UserResponse, status_code=201)
def register_user(
    user: UserPublicRegister,
    db: Session = Depends(get_db)
):
    """
    Registro público de nuevos usuarios.
    - Cualquier persona puede registrarse
    - El rol se deja como NULL (sin asignar)
    - Verifica que el email no exista
    """
    # Verificar si el email ya está registrado
    db_user = db.query(User).filter(User.email == user.email).first()
    
    if db_user:
        raise HTTPException(
            status_code=400, 
            detail="El email ya está registrado. Por favor, usa otro email o inicia sesión."
        )
    
    # Crear nuevo usuario con rol NULL (sin asignar)
    new_user = User(
        email=user.email,
        name=user.name,
        password=hash_password(user.password),
        role=None  # Rol sin asignar
    )
    
    # Guardar en la base de datos
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    return new_user


# ==========================================
# REGISTRO DE ADMIN (SOLO PARA ADMINISTRADORES)
# ==========================================
@router.post("/register-admin", response_model=UserResponse)
def register_admin(
    user: UserCreate,
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    """
    Registro de administradores (solo accesible para admins existentes).
    - Permite crear usuarios con rol admin, teacher o student
    """
    # Verificar si el email ya está registrado
    db_user = db.query(User).filter(User.email == user.email).first()
    
    if db_user:
        raise HTTPException(
            status_code=400, 
            detail="Email already registered"
        )
    
    # Crear nuevo usuario con el rol especificado
    new_user = User(
        email=user.email,
        name=user.name,
        password=hash_password(user.password),
        role=user.role.value
    )
    
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    return new_user


def _parse_excel_users(file: UploadFile) -> list[tuple[int, dict]]:
    allowed_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload a .xlsx file."
        )

    content = file.file.read()
    max_size = 5 * 1024 * 1024
    if len(content) > max_size:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="Excel file too large. Maximum size is 5 MB."
        )

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unable to read Excel file. Verify the file is a valid .xlsx workbook."
        )

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or not any(rows[0]):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file must have a header row with email and name."
        )

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    required_fields = {"email", "name"}
    valid_fields = {"email", "name", "password", "role"}
    missing = required_fields - set(headers)
    unknown = set(headers) - valid_fields

    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required columns: {', '.join(sorted(missing))}."
        )
    if unknown:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unknown columns found: {', '.join(sorted(unknown))}."
        )

    parsed = []
    for row_index, row in enumerate(rows[1:], start=2):
        if not any(cell is not None for cell in row):
            continue

        row_data = {}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            row_data[header] = row[col_index] if col_index < len(row) else None

        parsed.append((row_index, row_data))

    return parsed


@router.post("/upload", response_model=BulkUploadResult, status_code=status.HTTP_201_CREATED)
def upload_users(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    """
    Upload an Excel file (.xlsx) to create users in bulk.
    Columns must include: email, name, and optionally role.
    If a password is not provided, a temporary random password will be generated and emailed.
    """
    rows = _parse_excel_users(file)
    created_users = []
    errors = []

    for row_number, row_data in rows:
        row_data = {
            key: (str(value).strip() if value is not None else None)
            for key, value in row_data.items()
        }

        if row_data.get("email"):
            row_data["email"] = row_data["email"].lower()

        try:
            user_payload = UserBulkCreate(**row_data)
        except Exception as exc:
            errors.append({
                "row": row_number,
                "status": "error",
                "errors": getattr(exc, "errors", lambda: str(exc))(),
            })
            continue

        if db.query(User).filter(User.email == user_payload.email).first():
            errors.append({
                "row": row_number,
                "status": "error",
                "detail": "Email already registered",
            })
            continue

        temporary_password = user_payload.password or generate_random_password(12)
        new_user = User(
            email=user_payload.email,
            name=user_payload.name,
            password=hash_password(temporary_password),
            role=user_payload.role.value if user_payload.role else None,
            needs_password_change=True,
        )

        try:
            db.add(new_user)
            db.flush()
            send_temporary_password(new_user.email, temporary_password, new_user.name)
            db.commit()
            db.refresh(new_user)
            created_users.append(new_user)
        except Exception as exc:
            db.rollback()
            errors.append({
                "row": row_number,
                "status": "error",
                "detail": str(exc),
            })

    return BulkUploadResult(
        created_count=len(created_users),
        created=created_users,
        errors=errors,
    )


@router.get("/template", status_code=200)
def download_users_template(_: User = Depends(admin_required)):
    """
    Download an Excel template for bulk user upload.
    Columns: email (required), name (required), role (optional), password (optional).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "users_template"
    headers = ["email", "name", "role", "password"]
    ws.append(headers)
    ws.append(["student@example.com", "Student Name", "student", ""]) 

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = "excel_plantilla_users.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ----------------------------------------------------- 
# get all students (admin and teachers) - solo usuarios con rol student
@router.get("/students", response_model=list[UserResponse])
def get_students(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    return db.query(User).filter(User.role == "student").all()


# ------------------------------------------------
# get all teachers (admin and teachers) - solo usuarios con rol teacher
@router.get("/teachers", response_model=list[UserResponse])
def get_teachers(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    return db.query(User).filter(User.role == "teacher").all()


# ------------------------------------------------
# get all users (admin only) - incluye usuarios sin rol
@router.get("/", response_model=list[UserResponse])
def get_all_users(
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    return db.query(User).all()


# ------------------------------------------------
# get users without role (admin only) - NUEVO ENDPOINT
@router.get("/pending", response_model=list[UserResponse])
def get_pending_users(
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    """
    Obtener usuarios que aún no tienen rol asignado (pendientes de aprobación)
    """
    return db.query(User).filter(User.role.is_(None)).all()


# -----------------------------------------------------
# assign role to user (admin only) - NUEVO ENDPOINT
@router.put("/{user_id}/assign-role", response_model=UserResponse)
def assign_role(
    user_id: int,
    role: UserRole,
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    """
    Asignar un rol a un usuario que aún no tiene rol
    """
    user = db.query(User).filter(User.id == user_id).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    user.role = role.value
    db.commit()
    db.refresh(user)
    
    return user


# -----------------------------------------------------
# create user (admin only)
@router.post("/", response_model=UserResponse)
def create(
    user: UserCreate,
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    db_user = db.query(User).filter(User.email == user.email).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Email already registered")

    # Generar contraseña aleatoria (puedes usar la misma función que en upload)
    from ..services.email_service import generate_random_password, send_temporary_password
    temp_password = user.password or generate_random_password(12)
    hashed = hash_password(temp_password)

    new_user = User(
        email=user.email,
        name=user.name,
        password=hashed,
        role=user.role.value,
        needs_password_change=True   # Marcar que debe cambiar la contraseña
    )

    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    # Enviar correo con la contraseña temporal
    try:
        send_temporary_password(new_user.email, temp_password, new_user.name)
    except Exception as e:
        # Loggear error pero no fallar la creación
        print(f"Error enviando email a {new_user.email}: {e}")

    return new_user


# -----------------------------------------------------
# change user role (admin only)
@router.put("/{user_id}/role", response_model=UserResponse)
def change_role(
    user_id: int,
    role: UserRole,
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    user.role = role.value
    db.commit()
    db.refresh(user)

    return user


# ---------------------------------------------------------
# change own password (SIN verificar actual - simple)
@router.patch("/me/password", response_model=UserResponse)
def change_my_password(
    password_data: PasswordChange,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    user = db.query(User).filter(User.id == current_user.id).first()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    user.password = hash_password(password_data.password)
    user.needs_password_change = False
    db.commit()
    db.refresh(user)

    return user


# ---------------------------------------------------------
# NUEVO ENDPOINT: change own password (CON verificar actual)
@router.patch("/me/change-password", response_model=UserResponse)
def change_password_with_verification(
    password_data: PasswordChangeWithCurrent,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Cambiar la contraseña del usuario actual verificando la contraseña actual.
    Ideal para usuarios con contraseña temporal.
    """
    user = db.query(User).filter(User.id == current_user.id).first()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Verificar contraseña actual
    if not verify_password(password_data.current_password, user.password):
        raise HTTPException(status_code=401, detail="Contraseña actual incorrecta")

    # Actualizar contraseña
    user.password = hash_password(password_data.new_password)
    user.needs_password_change = False
    db.commit()
    db.refresh(user)

    return user


# ---------------------------------------------------------
# NUEVO ENDPOINT: Admin fuerza cambio de contraseña
@router.patch("/{user_id}/force-password-change", response_model=UserResponse)
def force_password_change(
    user_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    """
    Admin: Forzar a un usuario a cambiar su contraseña en el próximo login.
    """
    user = db.query(User).filter(User.id == user_id).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    user.needs_password_change = True
    db.commit()
    db.refresh(user)
    
    return user


# ---------------------------------------------------------
# get current authenticated user
@router.get("/me", response_model=UserResponse)
def get_my_profile(
    current_user: User = Depends(get_current_user)
):
    return current_user


# -------------------------------------------------------
# get user by id (admin and the user himself)
@router.get("/{user_id}", response_model=UserResponse)
def get_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # ADMIN -> acceso total
    if current_user.role == "admin":
        return user

    # USER -> sÃ­ mismo
    if current_user.id == user_id:
        return user

    # TEACHER -> solo alumnos de sus cursos
    if current_user.role == "teacher":
        
        is_student_in_teacher_course = db.query(EnrollmentDetail)\
            .join(
                Enrollment,
                Enrollment.id == EnrollmentDetail.enrollment_id
            )\
            .join(
                TeacherAssignment,
                TeacherAssignment.course_offering_id == EnrollmentDetail.offering_id
            )\
            .filter(
                Enrollment.student_id == user_id,
                TeacherAssignment.professor_id == current_user.id
            )\
            .first()

        if is_student_in_teacher_course:
            return user

    raise HTTPException(status_code=403, detail="Not authorized")

# ---------------------------------------------------------
# update user (admin only)
@router.put("/{user_id}", response_model=UserResponse)
def update_user(
    user_id: int,
    name: str,
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        raise HTTPException(404, "User not found")

    user.name = name
    db.commit()
    db.refresh(user)

    return user


# ---------------------------------------------------------
# delete user (admin only)
@router.delete("/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        raise HTTPException(404, "User not found")

    db.delete(user)
    db.commit()

    return {"message": "User deleted"}


# ---------------------------------------------------------
# get courses of a professor (admin and the professor himself)
@router.get("/{professor_id}/courses")
def get_professor_courses(
    professor_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ["admin", "teacher"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    professor = db.query(User).filter(User.id == professor_id).first()

    if not professor:
        raise HTTPException(status_code=404, detail="Professor not found")

    if professor.role != "teacher":
        raise HTTPException(status_code=400, detail="User is not a teacher")

    courses = [
        {
            "offering_id": ta.course_offering_id
        }
        for ta in professor.teacher_assignments
    ]

    return {
        "professor_id": professor.id,
        "professor_name": professor.name,
        "courses": courses
    }


# ---------------------------------------------------------
# patch user (admin and the user himself)
@router.patch("/{user_id}", response_model=UserResponse)
def patch_user(
    user_id: int,
    data: UserPatch,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        raise HTTPException(404, "User not found")

    if current_user.role != "admin" and current_user.id != user_id:
        raise HTTPException(403, "Not authorized")

    update_data = data.dict(exclude_unset=True)

    if "password" in update_data:
        update_data["password"] = hash_password(update_data["password"])

    for key, value in update_data.items():
        setattr(user, key, value)

    db.commit()
    db.refresh(user)

    return user


@router.get("/users/count")
def get_count_all_users(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Obtener el número total de usuarios - Solo administradores"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Se requieren permisos de administrador")

    total_users = db.query(User).count()
    return {"count": total_users}


# Contar usuarios por rol
@router.get("/users/count-by-role")
def get_users_by_role_count(
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    admins = db.query(User).filter(User.role == "admin").count()
    professors = db.query(User).filter(User.role == "teacher").count()
    students = db.query(User).filter(User.role == "student").count()
    return {"admins": admins, "professors": professors, "students": students}

# Buscar usuarios
@router.get("/users/search")
def search_users(
    q: str,
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    users = db.query(User).filter(
        (User.name.ilike(f"%{q}%")) | (User.email.ilike(f"%{q}%"))
    ).all()
    return {"users": users, "total": len(users)}

# Obtener sesiones de un usuario (admin)
@router.get("/{user_id}/sessions")
def get_user_sessions(
    user_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    from ..models.usersession_model import UserSession
    sessions = db.query(UserSession).filter(
        UserSession.user_id == user_id,
        UserSession.revoked == False
    ).all()
    return sessions

# Revocar una sesión específica (admin)
@router.delete("/{user_id}/sessions/{session_id}")
def revoke_user_session(
    user_id: int,
    session_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    from ..models.usersession_model import UserSession
    session = db.query(UserSession).filter(
        UserSession.id == session_id,
        UserSession.user_id == user_id
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    session.revoked = True
    db.commit()
    return {"message": "Session revoked"}

# Revocar todas las sesiones de un usuario (admin)
@router.delete("/{user_id}/sessions")
def revoke_all_user_sessions(
    user_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    from ..models.usersession_model import UserSession
    db.query(UserSession).filter(
        UserSession.user_id == user_id,
        UserSession.revoked == False
    ).update({"revoked": True}, synchronize_session=False)
    db.commit()
    return {"message": "All sessions revoked"}
