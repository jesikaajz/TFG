from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, status,Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
from openpyxl import load_workbook, Workbook
from sqlalchemy import or_ 
from typing import Optional,List

from ..bd.connection import get_db
from ..models.teacher_assignments_model import TeacherAssignment
from ..models.user_model import User
from ..models.course_offerings_model import CourseOffering
from ..models.subjects_model import Subject
from ..models.academic_year_model import AcademicYear
from ..models.enrollment_model import Enrollment
from ..models.enrollment_detail_model import EnrollmentDetail
from ..schemas.teacher_assignment_schema import (
    TeacherAssignmentCreate,
    TeacherAssignmentPatch,
    TeacherAssignmentResponse,
    BulkTeacherAssignmentResult,
    TeacherAssignmentBulkCreate
)
from ..dependencies.auth_dependencies import (
    admin_required,
    get_current_user
)

router = APIRouter(
    prefix="/teacher-assignments",
    tags=["Teacher Assignments"]
)


@router.get("/", response_model=list[TeacherAssignmentResponse])
def get_all(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user)
):
    return db.query(TeacherAssignment).all()

# teacher_assignment_routes.py

@router.get("/template", status_code=200)
def download_teacher_assignments_template(
    current_user: User = Depends(get_current_user)  # Cambiado de admin_required a get_current_user
):
    """
    Download an Excel template for bulk teacher assignment upload.
    Columns: professor_name, subject_name, academic_year_start, academic_year_end
    (professor_email can be used instead of professor_name).
    
    Permisos:
    - Admin: puede descargar
    - Teacher (tutor): puede descargar para preparar uploads
    """
    # Verificar que el usuario es admin o teacher
    if current_user.role not in ["admin", "teacher"]:
        raise HTTPException(403, "Solo administradores o profesores pueden descargar la plantilla")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "plantilla_asignaciones"
    headers = ["professor_name", "subject_name", "academic_year_start", "academic_year_end"]
    ws.append(headers)
    ws.append(["Dr. John Smith", "Algorithms", 2025, 2026])
    ws.append(["Prof. Jane Doe", "Data Structures", 2025, 2026])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = "plantilla_excel_teacher_assignments.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/{assignment_id}", response_model=TeacherAssignmentResponse)
def get_one(
    assignment_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user)
):
    assignment = db.query(TeacherAssignment).filter(
        TeacherAssignment.id == assignment_id
    ).first()

    if not assignment:
        raise HTTPException(404, "Assignment not found")

    return assignment


@router.post("/", response_model=TeacherAssignmentResponse)
def create(
    data: TeacherAssignmentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Verificar permisos: admin o tutor de la course offering
    if current_user.role == "admin":
        pass  # Admin puede crear cualquier asignación
    else:
        # Verificar si el usuario es tutor de esta course offering
        is_tutor = db.query(TeacherAssignment).filter(
            TeacherAssignment.course_offering_id == data.course_offering_id,
            TeacherAssignment.professor_id == current_user.id,
            TeacherAssignment.is_tutor == True
        ).first()
        
        if not is_tutor:
            raise HTTPException(403, "No tienes permiso para asignar profesores a este curso. Solo el tutor puede hacerlo.")
    
    # Verificar que la asignación no existe ya
    exists = db.query(TeacherAssignment).filter(
        TeacherAssignment.professor_id == data.professor_id,
        TeacherAssignment.course_offering_id == data.course_offering_id
    ).first()

    if exists:
        raise HTTPException(400, "Assignment already exists")

    assignment = TeacherAssignment(**data.model_dump())
    assignment.is_tutor = False  # Asegurar que no se crea como tutor

    db.add(assignment)
    db.commit()
    db.refresh(assignment)

    return assignment


@router.put("/{assignment_id}", response_model=TeacherAssignmentResponse)
def update(
    assignment_id: int,
    data: TeacherAssignmentCreate,
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    assignment = db.query(TeacherAssignment).filter(
        TeacherAssignment.id == assignment_id
    ).first()

    if not assignment:
        raise HTTPException(404, "Assignment not found")

    assignment.professor_id = data.professor_id
    assignment.course_offering_id = data.course_offering_id

    db.commit()
    db.refresh(assignment)

    return assignment


@router.patch("/{assignment_id}", response_model=TeacherAssignmentResponse)
def patch(
    assignment_id: int,
    data: TeacherAssignmentPatch,
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    assignment = db.query(TeacherAssignment).filter(
        TeacherAssignment.id == assignment_id
    ).first()

    if not assignment:
        raise HTTPException(404, "Assignment not found")

    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(assignment, key, value)

    db.commit()
    db.refresh(assignment)

    return assignment


@router.delete("/{assignment_id}")
def delete(
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    assignment = db.query(TeacherAssignment).filter(
        TeacherAssignment.id == assignment_id
    ).first()

    if not assignment:
        raise HTTPException(404, "Assignment not found")
    
    # Verificar permisos
    if current_user.role == "admin":
        # Admin puede eliminar cualquier asignación
        pass
    else:
        # Verificar si el usuario es tutor de esta course offering
        # Obtener la course offering de la asignación
        course_offering_id = assignment.course_offering_id
        
        is_tutor = db.query(TeacherAssignment).filter(
            TeacherAssignment.course_offering_id == course_offering_id,
            TeacherAssignment.professor_id == current_user.id,
            TeacherAssignment.is_tutor == True
        ).first()
        
        if not is_tutor:
            raise HTTPException(403, "No tienes permiso para eliminar asignaciones de este curso. Solo el tutor puede hacerlo.")
        
        # No permitir que el tutor se elimine a sí mismo
        if assignment.professor_id == current_user.id:
            raise HTTPException(400, "No puedes eliminar tu propia asignación como tutor")

    db.delete(assignment)
    db.commit()

    return {"message": "Assignment deleted"}


def _parse_excel_teacher_assignments(file: UploadFile) -> list[tuple[int, dict]]:
    allowed_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload a .xlsx file."
        )

    content = file.file.read()
    max_size = 5 * 1024 * 1024
    if len(content) > max_size:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="Excel file too large. Maximum size is 5 MB."
        )

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unable to read Excel file. Verify the file is a valid .xlsx workbook."
        )

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or not any(rows[0]):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file must have a header row with professor_name/professor_email, subject_name, academic_year_start, academic_year_end."
        )

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    valid_fields = {"professor_name", "professor_email", "subject_name", "academic_year_start", "academic_year_end"}
    has_professor = "professor_name" in headers or "professor_email" in headers
    has_subject = "subject_name" in headers
    has_year = "academic_year_start" in headers and "academic_year_end" in headers

    if not has_professor or not has_subject or not has_year:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel must include (professor_name or professor_email), subject_name, academic_year_start, and academic_year_end."
        )

    parsed = []
    for row_index, row in enumerate(rows[1:], start=2):
        if not any(cell is not None for cell in row):
            continue

        row_data = {}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            row_data[header] = row[col_index] if col_index < len(row) else None

        parsed.append((row_index, row_data))

    return parsed


@router.post("/upload", response_model=BulkTeacherAssignmentResult, status_code=status.HTTP_201_CREATED)
def upload_teacher_assignments(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)  # Cambiado de admin_required a get_current_user
):
    """
    Upload an Excel file (.xlsx) to create teacher assignments in bulk.
    Columns must include: (professor_name or professor_email), subject_name, academic_year_start, academic_year_end.
    
    Permisos:
    - Admin: puede subir asignaciones para cualquier course offering
    - Profesor TUTOR: solo puede subir asignaciones para course offerings donde es tutor
    """
    
    # Primero parsear el Excel para obtener los datos
    rows = _parse_excel_teacher_assignments(file)
    created_assignments = []
    errors = []

    # Si es profesor, obtener las offerings donde es tutor
    tutor_offering_ids = []
    is_admin = current_user.role == "admin"
    
    if not is_admin:
        if current_user.role != "teacher":
            raise HTTPException(403, "Solo administradores o profesores pueden subir asignaciones")
        
        # Obtener las course offerings donde el profesor es TUTOR
        tutor_assignments = db.query(TeacherAssignment).filter(
            TeacherAssignment.professor_id == current_user.id,
            TeacherAssignment.is_tutor == True
        ).all()
        
        tutor_offering_ids = [ta.course_offering_id for ta in tutor_assignments]
        
        if not tutor_offering_ids:
            raise HTTPException(403, "No eres tutor de ninguna asignatura. No puedes subir asignaciones.")

    for row_number, row_data in rows:
        row_data = {k: (str(v).strip() if v is not None else None) for k, v in row_data.items()}

        # Find professor by name or email
        professor_name = row_data.get("professor_name")
        professor_email = row_data.get("professor_email")
        professor = None

        if professor_email:
            professor_email = professor_email.lower()
            professor = db.query(User).filter(User.email == professor_email).first()
            if not professor:
                errors.append({"row": row_number, "status": "error", "detail": f"Professor with email '{professor_email}' not found"})
                continue
        elif professor_name:
            matches = db.query(User).filter(User.name == professor_name).all()
            if len(matches) == 1:
                professor = matches[0]
            elif len(matches) > 1:
                errors.append({"row": row_number, "status": "error", "detail": "Multiple professors found with this name. Provide email to disambiguate."})
                continue
            else:
                errors.append({"row": row_number, "status": "error", "detail": f"Professor with name '{professor_name}' not found"})
                continue
        else:
            errors.append({"row": row_number, "status": "error", "detail": "professor_name or professor_email required"})
            continue

        # No permitir que el tutor se asigne a sí mismo
        if not is_admin and professor.id == current_user.id:
            errors.append({"row": row_number, "status": "error", "detail": "No puedes asignarte a ti mismo como profesor"})
            continue

        # Find course offering by subject name + academic year
        subject_name = row_data.get("subject_name")
        if not subject_name:
            errors.append({"row": row_number, "status": "error", "detail": "subject_name cannot be empty"})
            continue

        try:
            start_year = int(float(row_data.get("academic_year_start", 0)))
            end_year = int(float(row_data.get("academic_year_end", 0)))
        except (ValueError, TypeError):
            errors.append({"row": row_number, "status": "error", "detail": "Invalid academic year values"})
            continue

        subject = db.query(Subject).filter(Subject.name == subject_name).first()
        if not subject:
            errors.append({"row": row_number, "status": "error", "detail": f"Subject '{subject_name}' not found"})
            continue

        academic_year = db.query(AcademicYear).filter(
            AcademicYear.start_year == start_year,
            AcademicYear.end_year == end_year
        ).first()
        if not academic_year:
            errors.append({"row": row_number, "status": "error", "detail": f"Academic year {start_year}-{end_year} not found"})
            continue

        course_offering = db.query(CourseOffering).filter(
            CourseOffering.subject_id == subject.id,
            CourseOffering.academic_year_id == academic_year.id
        ).first()
        if not course_offering:
            errors.append({"row": row_number, "status": "error", "detail": f"Course offering not found for subject '{subject_name}' in {start_year}-{end_year}"})
            continue

        # Si es profesor (no admin), verificar que la course offering está en sus tutorías
        if not is_admin and course_offering.id not in tutor_offering_ids:
            errors.append({
                "row": row_number, 
                "status": "error", 
                "detail": f"No tienes permiso para asignar profesores a '{subject_name}' ({start_year}-{end_year}) porque no eres tutor de esa asignatura"
            })
            continue

        # Check if assignment already exists
        existing = db.query(TeacherAssignment).filter(
            TeacherAssignment.professor_id == professor.id,
            TeacherAssignment.course_offering_id == course_offering.id
        ).first()
        if existing:
            errors.append({"row": row_number, "status": "error", "detail": "Teacher assignment already exists"})
            continue

        try:
            assignment = TeacherAssignment(
                professor_id=professor.id,
                course_offering_id=course_offering.id
            )
            db.add(assignment)
            db.commit()
            db.refresh(assignment)
            created_assignments.append(assignment)
        except Exception as exc:
            db.rollback()
            errors.append({"row": row_number, "status": "error", "detail": str(exc)})

    return BulkTeacherAssignmentResult(created_count=len(created_assignments), created=created_assignments, errors=errors)


# En teacher_assignment_routes.py - Agregar endpoints

# ==========================================
# ASIGNAR TUTOR A UNA COURSE OFFERING
# ==========================================
@router.patch("/{assignment_id}/set-tutor")
def set_tutor_status(
    assignment_id: int,
    is_tutor: bool,
    db: Session = Depends(get_db),
    current_user: User = Depends(admin_required)  # Solo admin puede asignar tutores
):
    """
    Asignar o quitar el rol de tutor a un profesor en una course offering.
    Los tutores pueden crear enrollments.
    Solo administradores pueden modificar este estado.
    """
    
    assignment = db.query(TeacherAssignment).filter(
        TeacherAssignment.id == assignment_id
    ).first()
    
    if not assignment:
        raise HTTPException(404, "Teacher assignment not found")
    
    assignment.is_tutor = is_tutor
    db.commit()
    db.refresh(assignment)
    
    return {
        "message": f"Tutor status updated to {is_tutor}",
        "assignment_id": assignment.id,
        "professor_id": assignment.professor_id,
        "course_offering_id": assignment.course_offering_id,
        "is_tutor": assignment.is_tutor
    }


# ==========================================
# OBTENER TUTORES DE UNA COURSE OFFERING
# ==========================================
@router.get("/course-offering/{course_offering_id}/tutors")
def get_tutors_by_course_offering(
    course_offering_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener los profesores que son tutores de una course offering.
    """
    
    assignments = db.query(TeacherAssignment).filter(
        TeacherAssignment.course_offering_id == course_offering_id,
        TeacherAssignment.is_tutor == True
    ).all()
    
    result = []
    for assignment in assignments:
        professor = db.query(User).filter(User.id == assignment.professor_id).first()
        if professor:
            result.append({
                "assignment_id": assignment.id,
                "professor_id": professor.id,
                "professor_name": professor.name,
                "professor_email": professor.email,
                "is_tutor": assignment.is_tutor
            })
    
    return result

# En teacher_assignment_routes.py

# ==========================================
# ASIGNAR PROFESOR COMO TUTOR (CREAR ASIGNACIÓN)
# ==========================================
@router.post("/assign-tutor")
def assign_tutor(
    professor_id: int,
    course_offering_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(admin_required)
):
    """
    Asignar un profesor como tutor de una course offering.
    Si ya existe una asignación normal, la actualiza a tutor.
    """
    
    # Verificar que el profesor existe
    professor = db.query(User).filter(
        User.id == professor_id,
        User.role == "teacher"
    ).first()
    
    if not professor:
        raise HTTPException(404, "Professor not found")
    
    # Verificar que la course offering existe
    offering = db.query(CourseOffering).filter(
        CourseOffering.id == course_offering_id
    ).first()
    
    if not offering:
        raise HTTPException(404, "Course offering not found")
    
    # Buscar o crear asignación
    assignment = db.query(TeacherAssignment).filter(
        TeacherAssignment.professor_id == professor_id,
        TeacherAssignment.course_offering_id == course_offering_id
    ).first()
    
    if not assignment:
        # Crear nueva asignación como tutor
        assignment = TeacherAssignment(
            professor_id=professor_id,
            course_offering_id=course_offering_id,
            is_tutor=True
        )
        db.add(assignment)
    else:
        # Actualizar a tutor
        assignment.is_tutor = True
    
    db.commit()
    db.refresh(assignment)
    
    return {
        "message": f"Professor {professor.name} is now tutor of the course offering",
        "assignment_id": assignment.id,
        "professor_id": professor_id,
        "course_offering_id": course_offering_id,
        "is_tutor": assignment.is_tutor
    }


# ==========================================
# OBTENER TODAS LAS ASIGNACIONES CON ROL DE TUTOR
# ==========================================
@router.get("/tutors")
def get_all_tutors(
    db: Session = Depends(get_db),
    current_user: User = Depends(admin_required)
):
    """
    Obtener todas las asignaciones donde el profesor es tutor.
    Solo administradores.
    """
    
    assignments = db.query(TeacherAssignment).filter(
        TeacherAssignment.is_tutor == True
    ).all()
    
    result = []
    for assignment in assignments:
        professor = db.query(User).filter(User.id == assignment.professor_id).first()
        offering = db.query(CourseOffering).filter(
            CourseOffering.id == assignment.course_offering_id
        ).first()
        
        result.append({
            "assignment_id": assignment.id,
            "professor_id": assignment.professor_id,
            "professor_name": professor.name if professor else None,
            "course_offering_id": assignment.course_offering_id,
            "subject_name": offering.subject.name if offering and offering.subject else None,
            "academic_year": f"{offering.academic_year.start_year}-{offering.academic_year.end_year}" if offering and offering.academic_year else None,
            "is_tutor": assignment.is_tutor
        })
    
    return result

# ==========================================
# OBTENER PROFESORES ASIGNADOS A UNA COURSE OFFERING (CON DETALLES)
# ==========================================
@router.get("/course-offering/{course_offering_id}/all")
def get_all_professors_by_course_offering(
    course_offering_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener TODOS los profesores asignados a una course offering (tanto tutores como profesores normales).
    
    Permisos:
    - Admin: puede ver cualquier course offering
    - Teacher: puede ver solo las course offerings donde está asignado (como tutor o profesor normal)
    - Student: puede ver los profesores de los cursos en los que está matriculado
    """
    
    # Verificar que la course offering existe
    offering = db.query(CourseOffering).filter(CourseOffering.id == course_offering_id).first()
    if not offering:
        raise HTTPException(404, "Course offering not found")
    
    # Verificar permisos
    if current_user.role == "admin":
        # Admin puede ver todo
        pass
    elif current_user.role == "teacher":
        # Verificar si el profesor está asignado a esta course offering
        is_authorized = db.query(TeacherAssignment).filter(
            TeacherAssignment.professor_id == current_user.id,
            TeacherAssignment.course_offering_id == course_offering_id
        ).first()
        
        if not is_authorized:
            raise HTTPException(403, "No tienes permiso para ver los profesores de esta oferta. Solo profesores asignados pueden ver esta información.")
    elif current_user.role == "student":
        # Estudiante: verificar que esté matriculado en este course offering
        # EnrollmentDetail tiene la relación directa con offering_id
        is_enrolled = db.query(EnrollmentDetail).join(
            Enrollment, EnrollmentDetail.enrollment_id == Enrollment.id
        ).filter(
            Enrollment.student_id == current_user.id,
            EnrollmentDetail.offering_id == course_offering_id
        ).first()
        
        if not is_enrolled:
            raise HTTPException(403, "No tienes permiso para ver los profesores de este curso. Solo estudiantes matriculados pueden ver esta información.")
    else:
        raise HTTPException(403, "No tienes permiso para ver esta información")
    
    # Obtener todas las asignaciones de esta course offering
    assignments = db.query(TeacherAssignment).filter(
        TeacherAssignment.course_offering_id == course_offering_id
    ).all()
    
    result = []
    for assignment in assignments:
        professor = db.query(User).filter(User.id == assignment.professor_id).first()
        if professor:
            result.append({
                "assignment_id": assignment.id,
                "professor_id": professor.id,
                "name": professor.name,
                "email": professor.email,
                "is_tutor": assignment.is_tutor,
                "is_current_user": professor.id == current_user.id
            })
    
    return {
        "course_offering_id": course_offering_id,
        "subject_name": offering.subject.name if offering.subject else None,
        "academic_year": f"{offering.academic_year.start_year}-{offering.academic_year.end_year}" if offering.academic_year else None,
        "total_assigned": len(result),
        "assignments": result
    }
# ==========================================
# OBTENER PROFESORES NO ASIGNADOS A UNA COURSE OFFERING
# ==========================================
@router.get("/course-offering/{course_offering_id}/available")
def get_available_professors_by_course_offering(
    course_offering_id: int,
    search: Optional[str] = Query(None, description="Filtrar por nombre o email"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener profesores NO asignados a una course offering específica.
    Útil para mostrar en un dropdown de asignación.
    
    Permisos:
    - Admin: puede ver cualquier course offering
    - Teacher: solo puede ver para course offerings donde es tutor
    """
    
    # Verificar que la course offering existe
    offering = db.query(CourseOffering).filter(CourseOffering.id == course_offering_id).first()
    if not offering:
        raise HTTPException(404, "Course offering not found")
    
    # Verificar permisos
    if current_user.role == "admin":
        # Admin puede ver todo
        pass
    elif current_user.role == "teacher":
        # Verificar si el profesor es tutor de esta course offering
        is_tutor = db.query(TeacherAssignment).filter(
            TeacherAssignment.professor_id == current_user.id,
            TeacherAssignment.course_offering_id == course_offering_id,
            TeacherAssignment.is_tutor == True
        ).first()
        
        if not is_tutor:
            raise HTTPException(403, "No tienes permiso para asignar profesores a esta oferta. Solo tutores pueden hacerlo.")
    else:
        raise HTTPException(403, "No tienes permiso para ver esta información")
    
    # Obtener IDs de profesores YA asignados a esta course offering
    assigned_professor_ids = db.query(TeacherAssignment.professor_id).filter(
        TeacherAssignment.course_offering_id == course_offering_id
    ).distinct().subquery()
    
    # Obtener profesores NO asignados
    query = db.query(User).filter(
        User.role == "teacher",
        User.id.notin_(assigned_professor_ids)
    )
    
    # Excluir al profesor actual (no puede asignarse a sí mismo)
    query = query.filter(User.id != current_user.id)
    
    # Filtrar por búsqueda si se proporciona
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                User.name.ilike(search_term),
                User.email.ilike(search_term)
            )
        )
    
    professors = query.order_by(User.name).all()
    
    return {
        "course_offering_id": course_offering_id,
        "total_available": len(professors),
        "professors": [
            {
                "id": p.id,
                "name": p.name,
                "email": p.email
            }
            for p in professors
        ]
    }


@router.post("/bulk", response_model=List[TeacherAssignmentResponse], status_code=status.HTTP_201_CREATED)
def create_assignments_bulk(
    data: TeacherAssignmentBulkCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Crear múltiples asignaciones de profesores en una sola petición.
    
    Permisos:
    - Admin: puede crear cualquier asignación
    - Profesor TUTOR: solo puede crear asignaciones para las course offerings donde es tutor
    
    Formato del JSON:
    {
        "assignments": [
            {"professor_id": 1, "course_offering_id": 10},
            {"professor_id": 2, "course_offering_id": 10},
            {"professor_id": 3, "course_offering_id": 11}
        ]
    }
    """
    
    if not data.assignments:
        raise HTTPException(400, "No assignments provided")
    
    created_assignments = []
    errors = []
    
    # Si es profesor, obtener las offerings donde es tutor (una sola vez)
    tutor_offering_ids = []
    is_admin = current_user.role == "admin"
    
    if not is_admin:
        if current_user.role != "teacher":
            raise HTTPException(403, "Solo administradores o profesores pueden crear asignaciones")
        
        # Obtener las course offerings donde el profesor es TUTOR
        tutor_assignments = db.query(TeacherAssignment).filter(
            TeacherAssignment.professor_id == current_user.id,
            TeacherAssignment.is_tutor == True
        ).all()
        
        tutor_offering_ids = [ta.course_offering_id for ta in tutor_assignments]
        
        if not tutor_offering_ids:
            raise HTTPException(403, "No eres tutor de ninguna asignatura. No puedes crear asignaciones.")
    
    for assignment_data in data.assignments:
        try:
            # Verificar que la course offering existe
            offering = db.query(CourseOffering).filter(
                CourseOffering.id == assignment_data.course_offering_id
            ).first()
            
            if not offering:
                errors.append({
                    "assignment": assignment_data.dict(),
                    "error": f"Course offering {assignment_data.course_offering_id} not found"
                })
                continue
            
            # Si es profesor (no admin), verificar que la course offering está en sus tutorías
            if not is_admin and assignment_data.course_offering_id not in tutor_offering_ids:
                errors.append({
                    "assignment": assignment_data.dict(),
                    "error": f"No tienes permiso para asignar profesores a la offering {assignment_data.course_offering_id} (no eres tutor)"
                })
                continue
            
            # Verificar que el profesor existe y es teacher
            professor = db.query(User).filter(
                User.id == assignment_data.professor_id,
                User.role == "teacher"
            ).first()
            
            if not professor:
                errors.append({
                    "assignment": assignment_data.dict(),
                    "error": f"Professor with id {assignment_data.professor_id} not found or is not a teacher"
                })
                continue
            
            # No permitir que el tutor se asigne a sí mismo
            if not is_admin and professor.id == current_user.id:
                errors.append({
                    "assignment": assignment_data.dict(),
                    "error": "No puedes asignarte a ti mismo como profesor"
                })
                continue
            
            # Verificar que la asignación no existe ya
            existing = db.query(TeacherAssignment).filter(
                TeacherAssignment.professor_id == assignment_data.professor_id,
                TeacherAssignment.course_offering_id == assignment_data.course_offering_id
            ).first()
            
            if existing:
                errors.append({
                    "assignment": assignment_data.dict(),
                    "error": "Assignment already exists"
                })
                continue
            
            # Crear la asignación
            assignment = TeacherAssignment(
                professor_id=assignment_data.professor_id,
                course_offering_id=assignment_data.course_offering_id,
                is_tutor=False  # Por defecto no es tutor
            )
            
            db.add(assignment)
            db.flush()  # Para obtener el ID sin commit aún
            created_assignments.append(assignment)
            
        except Exception as exc:
            errors.append({
                "assignment": assignment_data.dict(),
                "error": str(exc)
            })
    
    # Hacer commit de todas las asignaciones creadas
    if created_assignments:
        db.commit()
        for assignment in created_assignments:
            db.refresh(assignment)
    
    return created_assignments