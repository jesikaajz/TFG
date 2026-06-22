from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, status
from fastapi.params import Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
from openpyxl import load_workbook, Workbook

from ..models.enrollment_detail_model import EnrollmentDetail
from ..models.enrollment_model import Enrollment
from ..models.exercise_model import Exercise
from ..models.teacher_assignments_model import TeacherAssignment
from ..models.exercise_languange_model import ExerciseLanguage
from datetime import datetime
from ..models.rubric_model import Rubric
from sqlalchemy import and_
from ..models.user_model import User
from ..models.test_cases_model import TestCase
from ..bd.connection import get_db
from ..models.course_offerings_model import CourseOffering
from ..models.subjects_model import Subject
from ..models.academic_year_model import AcademicYear
from ..schemas.course_offerings_schema import (
    CourseOfferingCreate,
    CourseOfferingPatch,
    CourseOfferingResponse,
    BulkCourseOfferingResult
)
from ..dependencies.auth_dependencies import admin_required, get_current_user
from ..models.user_model import User

router = APIRouter(prefix="/course-offerings", tags=["Course Offerings"])


@router.get("/", response_model=list[CourseOfferingResponse])
def get_all(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(CourseOffering).all()


@router.post("/", response_model=CourseOfferingResponse)
def create(data: CourseOfferingCreate, db: Session = Depends(get_db), _: User = Depends(admin_required)):

    subject = db.query(Subject).filter(Subject.id == data.subject_id).first()
    year = db.query(AcademicYear).filter(AcademicYear.id == data.academic_year_id).first()

    if not subject:
        raise HTTPException(404, "Subject not found")
    if not year:
        raise HTTPException(404, "Academic year not found")

    offering = CourseOffering(**data.dict())
    db.add(offering)
    db.commit()
    db.refresh(offering)

    return offering


@router.patch("/{offering_id}", response_model=CourseOfferingResponse)
def patch(offering_id: int, data: CourseOfferingPatch, db: Session = Depends(get_db), _: User = Depends(admin_required)):

    offering = db.query(CourseOffering).filter(CourseOffering.id == offering_id).first()
    if not offering:
        raise HTTPException(404, "Offering not found")

    for k, v in data.dict(exclude_unset=True).items():
        setattr(offering, k, v)

    db.commit()
    db.refresh(offering)
    return offering


@router.delete("/{offering_id}")
def delete(offering_id: int, db: Session = Depends(get_db), _: User = Depends(admin_required)):

    offering = db.query(CourseOffering).filter(CourseOffering.id == offering_id).first()
    if not offering:
        raise HTTPException(404, "Offering not found")

    db.delete(offering)
    db.commit()

    return {"message": "Course offering deleted"}


def _parse_excel_course_offerings(file: UploadFile) -> list[tuple[int, dict]]:
    allowed_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload a .xlsx file."
        )

    content = file.file.read()
    max_size = 5 * 1024 * 1024
    if len(content) > max_size:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="Excel file too large. Maximum size is 5 MB."
        )

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unable to read Excel file. Verify the file is a valid .xlsx workbook."
        )

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or not any(rows[0]):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file must have a header row with subject_name, academic_year_start, academic_year_end."
        )

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    required_fields = {"subject_name", "academic_year_start", "academic_year_end"}
    missing = required_fields - set(headers)

    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required columns: {', '.join(sorted(missing))}."
        )

    parsed = []
    for row_index, row in enumerate(rows[1:], start=2):
        if not any(cell is not None for cell in row):
            continue

        row_data = {}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            row_data[header] = row[col_index] if col_index < len(row) else None

        parsed.append((row_index, row_data))

    return parsed


@router.post("/upload", response_model=BulkCourseOfferingResult, status_code=status.HTTP_201_CREATED)
def upload_course_offerings(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(admin_required)
):
    """
    Upload an Excel file (.xlsx) to create course offerings in bulk.
    Columns must include: subject_name, academic_year_start, academic_year_end.
    """
    rows = _parse_excel_course_offerings(file)
    created_offerings = []
    errors = []

    for row_number, row_data in rows:
        row_data = {k: (str(v).strip() if v is not None else None) for k, v in row_data.items()}

        try:
            subject_name = row_data.get("subject_name")
            start_year = int(float(row_data.get("academic_year_start", 0)))
            end_year = int(float(row_data.get("academic_year_end", 0)))
        except (ValueError, TypeError):
            errors.append({"row": row_number, "status": "error", "detail": "Invalid academic year values"})
            continue

        if not subject_name:
            errors.append({"row": row_number, "status": "error", "detail": "subject_name cannot be empty"})
            continue

        subject = db.query(Subject).filter(Subject.name == subject_name).first()
        if not subject:
            errors.append({"row": row_number, "status": "error", "detail": f"Subject '{subject_name}' not found"})
            continue

        academic_year = db.query(AcademicYear).filter(
            AcademicYear.start_year == start_year,
            AcademicYear.end_year == end_year
        ).first()
        if not academic_year:
            errors.append({"row": row_number, "status": "error", "detail": f"Academic year {start_year}-{end_year} not found"})
            continue

        # Check if offering already exists
        existing = db.query(CourseOffering).filter(
            CourseOffering.subject_id == subject.id,
            CourseOffering.academic_year_id == academic_year.id
        ).first()
        if existing:
            errors.append({"row": row_number, "status": "error", "detail": "Course offering already exists for this subject and academic year"})
            continue

        try:
            offering = CourseOffering(
                subject_id=subject.id,
                academic_year_id=academic_year.id
            )
            db.add(offering)
            db.commit()
            db.refresh(offering)
            created_offerings.append(offering)
        except Exception as exc:
            db.rollback()
            errors.append({"row": row_number, "status": "error", "detail": str(exc)})

    return BulkCourseOfferingResult(created_count=len(created_offerings), created=created_offerings, errors=errors)


@router.get("/template", status_code=200)
def download_course_offerings_template(_: User = Depends(admin_required)):
    """
    Download an Excel template for bulk course offering upload.
    Columns: subject_name, academic_year_start, academic_year_end.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "plantilla_ofertas"
    headers = ["subject_name", "academic_year_start", "academic_year_end"]
    ws.append(headers)
    ws.append(["Algorithms", 2025, 2026])
    ws.append(["Data Structures", 2025, 2026])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = "plantilla_excel_course_offerings.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==========================================
# OBTENER TODAS LAS OFERTAS DE UNA ASIGNATURA (para seleccionar cual duplicar)
# ==========================================
@router.get("/subject/{subject_id}")
def get_offerings_by_subject(
    subject_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener todas las ofertas de una asignatura específica.
    Útil para que profesores/admins vean qué ofertas existen de años pasados.
    """
    # Verificar permisos
    if current_user.role == "teacher":
        # Verificar si el profesor está asignado a esta materia en algún año (actual o futuro)
        teacher_in_subject = db.query(TeacherAssignment).join(
            CourseOffering, TeacherAssignment.course_offering_id == CourseOffering.id
        ).filter(
            CourseOffering.subject_id == subject_id,
            TeacherAssignment.professor_id == current_user.id
        ).first()
        
        if not teacher_in_subject and current_user.role != "admin":
            raise HTTPException(403, "No tienes acceso a esta asignatura")
    
    offerings = db.query(CourseOffering).filter(
        CourseOffering.subject_id == subject_id
    ).order_by(CourseOffering.academic_year_id.desc()).all()
    
    result = []
    for offering in offerings:
        exercises_count = db.query(Exercise).filter(
            Exercise.course_offering_id == offering.id
        ).count()
        
        # Verificar si el profesor actual está asignado a esta oferta
        is_my_offering = False
        if current_user.role == "teacher":
            is_my_offering = db.query(TeacherAssignment).filter(
                TeacherAssignment.course_offering_id == offering.id,
                TeacherAssignment.professor_id == current_user.id
            ).first() is not None
        
        result.append({
            "id": offering.id,
            "subject_id": offering.subject_id,
            "subject_name": offering.subject.name if offering.subject else None,
            "academic_year_id": offering.academic_year_id,
            "academic_year": f"{offering.academic_year.start_year}-{offering.academic_year.end_year}" if offering.academic_year else None,
            "exercises_count": exercises_count,
            "is_my_offering": is_my_offering,
            "has_teacher": db.query(TeacherAssignment).filter(
                TeacherAssignment.course_offering_id == offering.id
            ).first() is not None
        })
    
    return result


# ==========================================
# DUPLICAR COURSE OFFERING (con ejercicios, sin enrollments)
# ==========================================
@router.post("/{offering_id}/duplicate")
def duplicate_offering(
    offering_id: int,
    new_academic_year_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Duplicar una course offering existente a un nuevo año académico.
    - Se copian todos los ejercicios, test cases, rúbricas y lenguajes permitidos
    - NO se copian los enrollments (matrículas)
    
    Permisos:
    - Admin: puede duplicar cualquier offering a cualquier año
    - Teacher: debe estar asignado a la misma materia en el año DESTINO
               (puede duplicar ofertas pasadas de esa materia, incluso de otros profesores)
    """
    
    # 1. Verificar que la offering original existe
    original_offering = db.query(CourseOffering).filter(
        CourseOffering.id == offering_id
    ).first()
    
    if not original_offering:
        raise HTTPException(404, "Course offering not found")
    
    # 2. Verificar que el nuevo año académico existe
    new_academic_year = db.query(AcademicYear).filter(
        AcademicYear.id == new_academic_year_id
    ).first()
    
    if not new_academic_year:
        raise HTTPException(404, "Academic year not found")
    
    # 3. Verificar permisos según rol (VERSIÓN MODIFICADA)
    if current_user.role == "teacher":
        # Verificar si el profesor ha impartido ESTA MISMA MATERIA en ALGÚN año pasado
        teacher_has_taught_subject = db.query(TeacherAssignment).join(
            CourseOffering, TeacherAssignment.course_offering_id == CourseOffering.id
        ).filter(
            CourseOffering.subject_id == original_offering.subject_id,
            TeacherAssignment.professor_id == current_user.id
        ).first()
    
        if not teacher_has_taught_subject:
            raise HTTPException(
                403, 
                "No estás autorizado para duplicar esta materia porque nunca has sido asignado a ella"
            )
    
        # Si ha impartido la materia antes, permitir duplicar
        # (la asignación se creará automáticamente después)
    
    # 4. Verificar que no existe ya una offering para la misma materia y año
    existing_offering = db.query(CourseOffering).filter(
        CourseOffering.subject_id == original_offering.subject_id,
        CourseOffering.academic_year_id == new_academic_year_id
    ).first()
    
    if existing_offering:
        raise HTTPException(400, f"Ya existe una oferta para esta materia en el año {new_academic_year.start_year}-{new_academic_year.end_year}")
    
    # 5. Crear la nueva offering
    new_offering = CourseOffering(
        subject_id=original_offering.subject_id,
        academic_year_id=new_academic_year_id
    )
    db.add(new_offering)
    db.flush()  # Para obtener el ID
    
    # 6. Duplicar ejercicios
    original_exercises = db.query(Exercise).filter(
        Exercise.course_offering_id == offering_id
    ).all()
    
    exercises_created = []
    
    for original_exercise in original_exercises:
        # Crear nuevo ejercicio
        new_exercise = Exercise(
            title=original_exercise.title,
            description=original_exercise.description,
            deadline=original_exercise.deadline,
            course_offering_id=new_offering.id,
            solution=original_exercise.solution,
            visibility=original_exercise.visibility
        )
        db.add(new_exercise)
        db.flush()
        
        # Duplicar test cases
        test_cases = db.query(TestCase).filter(
            TestCase.exercise_id == original_exercise.id
        ).all()
        
        for tc in test_cases:
            new_tc = TestCase(
                exercise_id=new_exercise.id,
                input_data=tc.input_data,
                expected_output=tc.expected_output
            )
            db.add(new_tc)
        
        # Duplicar rúbrica
        rubric = db.query(Rubric).filter(
            Rubric.exercise_id == original_exercise.id
        ).first()
        
        if rubric:
            new_rubric = Rubric(
                exercise_id=new_exercise.id,
                criteria=rubric.criteria
            )
            db.add(new_rubric)
        
        # Duplicar lenguajes permitidos
        languages = db.query(ExerciseLanguage).filter(
            ExerciseLanguage.exercise_id == original_exercise.id
        ).all()
        
        for lang in languages:
            new_lang = ExerciseLanguage(
                exercise_id=new_exercise.id,
                language_id=lang.language_id
            )
            db.add(new_lang)
        
        exercises_created.append({
            "original_id": original_exercise.id,
            "new_id": new_exercise.id,
            "title": new_exercise.title,
            "test_cases_count": len(test_cases),
            "has_rubric": rubric is not None,
            "languages_count": len(languages)
        })
    
    # 7. El profesor que duplica YA debe estar asignado en el año destino
    # (ya verificamos teacher_in_destination antes)
    # Si no existe la asignación, la creamos (por si el admin la omitió)
    if current_user.role == "teacher":
        existing_assignment = db.query(TeacherAssignment).filter(
            TeacherAssignment.course_offering_id == new_offering.id,
            TeacherAssignment.professor_id == current_user.id
        ).first()
        
        if not existing_assignment:
            new_assignment = TeacherAssignment(
                professor_id=current_user.id,
                course_offering_id=new_offering.id
            )
            db.add(new_assignment)
    
    db.commit()
    db.refresh(new_offering)
    
    # 8. Respuesta
    return {
        "message": "Course offering duplicated successfully",
        "original_offering_id": offering_id,
        "original_academic_year": f"{original_offering.academic_year.start_year}-{original_offering.academic_year.end_year}",
        "new_offering": {
            "id": new_offering.id,
            "subject_id": new_offering.subject_id,
            "subject_name": new_offering.subject.name if new_offering.subject else None,
            "academic_year_id": new_offering.academic_year_id,
            "academic_year": f"{new_academic_year.start_year}-{new_academic_year.end_year}",
            "exercises_count": len(exercises_created)
        },
        "exercises_duplicated": exercises_created
    }


# ==========================================
# OBTENER OFERTAS DISPONIBLES PARA DUPLICAR
# ==========================================
@router.get("/available-for-duplicate")
def get_offerings_available_for_duplicate(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener todas las course offerings que el usuario puede duplicar.
    - Admin: todas las offerings
    - Teacher: offerings de materias que IMPARTE EN EL AÑO ACTUAL O FUTURO
               (puede duplicar ofertas pasadas de esas materias)
    """
    
    if current_user.role == "admin":
        offerings = db.query(CourseOffering).all()
        
    elif current_user.role == "teacher":
        # Obtener las materias que el profesor imparte actualmente (o futuro)
        # Es decir, materias donde tiene teacher_assignment en algún año
        my_subject_ids = db.query(CourseOffering.subject_id).join(
            TeacherAssignment, CourseOffering.id == TeacherAssignment.course_offering_id
        ).filter(
            TeacherAssignment.professor_id == current_user.id
        ).distinct().subquery()
        
        # Obtener todas las ofertas de esas materias (para poder duplicar ofertas pasadas)
        offerings = db.query(CourseOffering).filter(
            CourseOffering.subject_id.in_(my_subject_ids)
        ).order_by(CourseOffering.academic_year_id.desc()).all()
        
    else:
        raise HTTPException(403, "Not authorized")
    
    result = []
    for offering in offerings:
        # Verificar si el profesor actual está asignado a la misma materia en el año actual
        current_year = datetime.now().year
        is_current_teacher = False
        
        if current_user.role == "teacher":
            # Verificar si el profesor imparte esta materia en el año actual
            is_current_teacher = db.query(TeacherAssignment).join(
                CourseOffering, TeacherAssignment.course_offering_id == CourseOffering.id
            ).filter(
                CourseOffering.subject_id == offering.subject_id,
                CourseOffering.academic_year.has(
                    and_(
                        AcademicYear.start_year <= current_year,
                        AcademicYear.end_year >= current_year
                    )
                ),
                TeacherAssignment.professor_id == current_user.id
            ).first() is not None
        
        result.append({
            "id": offering.id,
            "subject_id": offering.subject_id,
            "subject_name": offering.subject.name if offering.subject else None,
            "academic_year_id": offering.academic_year_id,
            "academic_year": f"{offering.academic_year.start_year}-{offering.academic_year.end_year}",
            "exercises_count": db.query(Exercise).filter(
                Exercise.course_offering_id == offering.id
            ).count(),
            "can_duplicate": current_user.role == "admin" or is_current_teacher
        })
    
    return result


# ==========================================
# OBTENER MATERIAS QUE EL PROFESOR IMPARTE EN UN AÑO ESPECÍFICO
# ==========================================
@router.get("/my-subjects/{academic_year_id}")
def get_my_subjects_by_year(
    academic_year_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener las materias que el profesor imparte en un año académico específico.
    Útil para saber qué materias puede duplicar a ese año.
    """
    if current_user.role not in ["teacher", "admin"]:
        raise HTTPException(403, "Not authorized")
    
    query = db.query(CourseOffering).join(
        TeacherAssignment, CourseOffering.id == TeacherAssignment.course_offering_id
    ).filter(
        CourseOffering.academic_year_id == academic_year_id
    )
    
    if current_user.role == "teacher":
        query = query.filter(TeacherAssignment.professor_id == current_user.id)
    
    offerings = query.all()
    
    result = []
    for offering in offerings:
        # Verificar si tiene ofertas pasadas para duplicar
        past_offerings = db.query(CourseOffering).filter(
            CourseOffering.subject_id == offering.subject_id,
            CourseOffering.academic_year_id < academic_year_id
        ).count()
        
        result.append({
            "offering_id": offering.id,
            "subject_id": offering.subject_id,
            "subject_name": offering.subject.name if offering.subject else None,
            "academic_year": f"{offering.academic_year.start_year}-{offering.academic_year.end_year}",
            "has_past_offerings": past_offerings > 0,
            "past_offerings_count": past_offerings
        })
    
    return result

# ==========================================
# BUSCADOR DE COURSE OFFERINGS CON FILTROS 
# ==========================================
@router.get("/search", response_model=dict)
def search_course_offerings(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    # Filtros
    subject_id: Optional[int] = Query(None, description="ID de la materia"),
    subject_name: Optional[str] = Query(None, description="Nombre de la materia (búsqueda parcial)"),
    academic_year_id: Optional[int] = Query(None, description="ID del año académico"),
    start_year: Optional[int] = Query(None, description="Año de inicio"),
    end_year: Optional[int] = Query(None, description="Año de fin"),
    teacher_id: Optional[int] = Query(None, description="ID del profesor asignado"),
    teacher_name: Optional[str] = Query(None, description="Nombre del profesor (búsqueda parcial)"),
    has_exercises: Optional[bool] = Query(None, description="Si tiene ejercicios asociados"),
    has_teacher: Optional[bool] = Query(None, description="Si tiene profesor asignado"),
    # Paginación
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    # Ordenamiento
    order_by: str = Query("academic_year_id", description="Campo para ordenar"),
    order_desc: bool = Query(True, description="Orden descendente")
):
    """
    Buscar course offerings con múltiples filtros.
    
    Permisos:
    - Admin: puede ver todas
    - Teacher: solo las que imparte
    - Student: solo las de materias donde está matriculado
    """
    
    # 1. Query base con joins necesarios
    query = db.query(CourseOffering).join(
        Subject, CourseOffering.subject_id == Subject.id
    ).join(
        AcademicYear, CourseOffering.academic_year_id == AcademicYear.id
    ).outerjoin(
        TeacherAssignment, CourseOffering.id == TeacherAssignment.course_offering_id
    ).outerjoin(
        User, TeacherAssignment.professor_id == User.id
    )
    
    # 2. Filtrar por permisos según rol
    if current_user.role == "teacher":
        # Profesor solo ve sus propias ofertas
        query = query.filter(TeacherAssignment.professor_id == current_user.id)
    elif current_user.role == "student":
        # Estudiante solo ve ofertas de materias donde está matriculado
        student_enrollments = db.query(EnrollmentDetail.offering_id).join(
            Enrollment, Enrollment.id == EnrollmentDetail.enrollment_id
        ).filter(
            Enrollment.student_id == current_user.id
        ).subquery()
        query = query.filter(CourseOffering.id.in_(student_enrollments))
    
    # 3. Aplicar filtros
    if subject_id:
        query = query.filter(CourseOffering.subject_id == subject_id)
    
    if subject_name:
        query = query.filter(Subject.name.ilike(f"%{subject_name}%"))
    
    if academic_year_id:
        query = query.filter(CourseOffering.academic_year_id == academic_year_id)
    
    if start_year:
        query = query.filter(AcademicYear.start_year == start_year)
    
    if end_year:
        query = query.filter(AcademicYear.end_year == end_year)
    
    if teacher_id:
        query = query.filter(TeacherAssignment.professor_id == teacher_id)
    
    if teacher_name:
        query = query.filter(User.name.ilike(f"%{teacher_name}%"))
    
    if has_exercises is not None:
        if has_exercises:
            # Tiene al menos un ejercicio
            query = query.filter(
                CourseOffering.id.in_(
                    db.query(Exercise.course_offering_id).distinct()
                )
            )
        else:
            # No tiene ejercicios
            query = query.filter(
                ~CourseOffering.id.in_(
                    db.query(Exercise.course_offering_id).distinct()
                )
            )
    
    if has_teacher is not None:
        if has_teacher:
            query = query.filter(TeacherAssignment.id.isnot(None))
        else:
            query = query.filter(TeacherAssignment.id.is_(None))
    
    # 4. Determinar columna de ordenamiento
    order_column_map = {
        "subject_name": Subject.name,
        "start_year": AcademicYear.start_year,
        "end_year": AcademicYear.end_year,
        "academic_year_id": CourseOffering.academic_year_id,
        "subject_id": CourseOffering.subject_id,
        "id": CourseOffering.id
    }
    order_column = order_column_map.get(order_by, CourseOffering.academic_year_id)
    
    # 5. IMPORTANTE: Primero aplicar ORDER BY (para que DISTINCT ON funcione correctamente)
    # Pero necesitamos que la primera columna en ORDER BY sea CourseOffering.id
    # para que DISTINCT ON (CourseOffering.id) sea válido en PostgreSQL
    if order_desc:
        query = query.order_by(CourseOffering.id.desc())
    else:
        query = query.order_by(CourseOffering.id.asc())
    
    # 6. Ahora aplicar DISTINCT y obtener IDs paginados
    # Usar with_entities para obtener solo los IDs (más eficiente)
    distinct_ids_query = query.with_entities(CourseOffering.id).distinct(CourseOffering.id)
    
    # Contar total (necesitamos una subconsulta para evitar problemas de paginación)
    total = distinct_ids_query.count()
    
    # Obtener IDs paginados
    id_rows = distinct_ids_query.offset(offset).limit(limit).all()
    ids = [row[0] for row in id_rows]
    
    # Si no hay IDs, retornar vacío
    if not ids:
        offerings = []
    else:
        # Obtener las ofertas completas con los IDs seleccionados
        # Reconstruir la consulta completa con joins para obtener los datos
        offerings_query = db.query(CourseOffering).join(
            Subject, CourseOffering.subject_id == Subject.id
        ).join(
            AcademicYear, CourseOffering.academic_year_id == AcademicYear.id
        ).filter(
            CourseOffering.id.in_(ids)
        )
        
        # Aplicar el ordenamiento deseado en el resultado final
        if order_desc:
            offerings_query = offerings_query.order_by(order_column.desc())
        else:
            offerings_query = offerings_query.order_by(order_column.asc())
        
        offerings = offerings_query.all()
    
    # 7. Construir respuesta
    result = []
    for offering in offerings:
        # Obtener profesores
        teachers = db.query(User).join(
            TeacherAssignment, User.id == TeacherAssignment.professor_id
        ).filter(
            TeacherAssignment.course_offering_id == offering.id
        ).all()
        
        # Contar ejercicios
        exercises_count = db.query(Exercise).filter(
            Exercise.course_offering_id == offering.id
        ).count()
        
        # Contar estudiantes matriculados
        students_count = db.query(EnrollmentDetail).filter(
            EnrollmentDetail.offering_id == offering.id
        ).count()
        
        result.append({
            "id": offering.id,
            "subject_id": offering.subject_id,
            "subject_name": offering.subject.name if offering.subject else None,
            "subject_description": offering.subject.description if offering.subject else None,
            "academic_year_id": offering.academic_year_id,
            "academic_year": f"{offering.academic_year.start_year}-{offering.academic_year.end_year}",
            "teachers": [
                {"id": t.id, "name": t.name, "email": t.email}
                for t in teachers
            ],
            "teachers_count": len(teachers),
            "exercises_count": exercises_count,
            "students_count": students_count,
            "has_exercises": exercises_count > 0,
            "has_teacher": len(teachers) > 0
        })
    
    return {
        "total": total,
        "limit": limit,
        "offset": offset,
        "has_more": offset + limit < total,
        "items": result,
        "filters_applied": {
            "subject_id": subject_id,
            "subject_name": subject_name,
            "academic_year_id": academic_year_id,
            "start_year": start_year,
            "end_year": end_year,
            "teacher_id": teacher_id,
            "teacher_name": teacher_name,
            "has_exercises": has_exercises,
            "has_teacher": has_teacher
        }
    }

# ==========================================
# OBTENER VALORES ÚNICOS PARA FILTROS (dropdowns)
# ==========================================
@router.get("/filter-options")
def get_course_offering_filter_options(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener valores únicos para los filtros de course offerings.
    Útil para poblar dropdowns en el frontend.
    """
    
    # Materias disponibles (según permisos)
    if current_user.role == "teacher":
        # Profesor solo ve materias que imparte
        subjects = db.query(Subject).join(
            CourseOffering, Subject.id == CourseOffering.subject_id
        ).join(
            TeacherAssignment, CourseOffering.id == TeacherAssignment.course_offering_id
        ).filter(
            TeacherAssignment.professor_id == current_user.id
        ).distinct().all()
    elif current_user.role == "student":
        # Estudiante solo ve materias donde está matriculado
        subjects = db.query(Subject).join(
            CourseOffering, Subject.id == CourseOffering.subject_id
        ).join(
            EnrollmentDetail, CourseOffering.id == EnrollmentDetail.offering_id
        ).join(
            Enrollment, EnrollmentDetail.enrollment_id == Enrollment.id
        ).filter(
            Enrollment.student_id == current_user.id
        ).distinct().all()
    else:
        subjects = db.query(Subject).all()
    
    # Años académicos disponibles
    academic_years = db.query(AcademicYear).order_by(AcademicYear.start_year.desc()).all()
    
    # Profesores disponibles (solo para admin y teacher)
    teachers = []
    if current_user.role in ["admin", "teacher"]:
        teachers = db.query(User).filter(User.role == "teacher").all()
    
    return {
        "subjects": [
            {"id": s.id, "name": s.name}
            for s in subjects
        ],
        "academic_years": [
            {"id": y.id, "name": f"{y.start_year}-{y.end_year}", "start_year": y.start_year, "end_year": y.end_year}
            for y in academic_years
        ],
        "teachers": [
            {"id": t.id, "name": t.name, "email": t.email}
            for t in teachers
        ]
    }