# enrollments_routes.py - Versión CORREGIDA con orden de rutas adecuado
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, status, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import date
from typing import List, Optional
from io import BytesIO
from openpyxl import load_workbook, Workbook
from sqlalchemy import or_
from ..bd.connection import get_db
from ..models.enrollment_model import Enrollment
from ..models.enrollment_detail_model import EnrollmentDetail
from ..models.course_offerings_model import CourseOffering
from ..models.subjects_model import Subject
from ..models.academic_year_model import AcademicYear
from ..models.user_model import User
from ..models.teacher_assignments_model import TeacherAssignment
from ..schemas.enrollment_schema import (
    EnrollmentCreate,
    EnrollmentResponse,
    EnrollmentUpdate,
    EnrollmentWithCoursesResponse,
    CourseInfo,
    EnrollmentBulkCreate,
    BulkEnrollmentResult,
)
from ..dependencies.auth_dependencies import admin_required, get_current_user
import os
import re
import secrets
import time
from ..services.email_service import generate_random_password, send_temporary_password
from ..security.jwt_handler import hash_password

router = APIRouter(prefix="/enrollments", tags=["Enrollments"])


# ==========================================
# FUNCIÓN AUXILIAR: Verificar si un profesor es tutor de alguna course offering
# ==========================================
def is_teacher_tutor_of_offerings(db: Session, teacher_id: int, offering_ids: List[int]) -> bool:
    """
    Verifica si un profesor es TUTOR (is_tutor=True) de al menos una de las course offerings.
    """
    if not offering_ids:
        return False
    
    tutor_assignment = db.query(TeacherAssignment).filter(
        TeacherAssignment.professor_id == teacher_id,
        TeacherAssignment.course_offering_id.in_(offering_ids),
        TeacherAssignment.is_tutor == True
    ).first()
    
    return tutor_assignment is not None


def is_teacher_tutor_of_offering(db: Session, teacher_id: int, offering_id: int) -> bool:
    """
    Verifica si un profesor es TUTOR (is_tutor=True) de una course offering específica.
    """
    tutor_assignment = db.query(TeacherAssignment).filter(
        TeacherAssignment.professor_id == teacher_id,
        TeacherAssignment.course_offering_id == offering_id,
        TeacherAssignment.is_tutor == True
    ).first()
    
    return tutor_assignment is not None


def get_tutor_offering_ids(db: Session, teacher_id: int) -> List[int]:
    """
    Obtener todos los IDs de course offerings donde un profesor es tutor.
    """
    assignments = db.query(TeacherAssignment.course_offering_id).filter(
        TeacherAssignment.professor_id == teacher_id,
        TeacherAssignment.is_tutor == True
    ).all()
    
    return [a[0] for a in assignments]


def _parse_excel_enrollments(file: UploadFile) -> list[tuple[int, dict]]:
    allowed_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload a .xlsx file."
        )

    content = file.file.read()
    max_size = 5 * 1024 * 1024
    if len(content) > max_size:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="Excel file too large. Maximum size is 5 MB."
        )

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unable to read Excel file. Verify the file is a valid .xlsx workbook."
        )

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or not any(rows[0]):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file must have a header row with student_id and academic_year_id."
        )

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    required_fields = {"student_id", "academic_year_id"}
    valid_fields = {"student_id", "academic_year_id", "offering_ids"}
    missing = required_fields - set(headers)
    unknown = set(headers) - valid_fields

    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required columns: {', '.join(sorted(missing))}."
        )
    if unknown:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unknown columns found: {', '.join(sorted(unknown))}."
        )

    parsed = []
    for row_index, row in enumerate(rows[1:], start=2):
        if not any(cell is not None for cell in row):
            continue

        row_data = {}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            row_data[header] = row[col_index] if col_index < len(row) else None

        parsed.append((row_index, row_data))

    return parsed


def _parse_excel_enrollments_with_users(file: UploadFile) -> list[tuple[int, dict]]:
    allowed_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload a .xlsx file."
        )

    content = file.file.read()
    max_size = 5 * 1024 * 1024
    if len(content) > max_size:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="Excel file too large. Maximum size is 5 MB."
        )

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unable to read Excel file. Verify the file is a valid .xlsx workbook."
        )

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or not any(rows[0]):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file must have a header row with student identifiers and academic_year_id."
        )

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    valid_fields = {"student_id", "email", "name", "role", "password", "academic_year_id", "academic_year_start", "academic_year_end", "offering_ids", "offering_subject", "offering_academic_year_start", "offering_academic_year_end"}
    has_academic_year = "academic_year_id" in headers or ("academic_year_start" in headers and "academic_year_end" in headers)
    has_student = "student_id" in headers or "email" in headers or "name" in headers
    if not has_academic_year:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel must include 'academic_year_id' or both 'academic_year_start' and 'academic_year_end' columns."
        )
    if not has_student:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel must include 'student_id', 'email', or 'name' to identify students."
        )

    unknown = set(headers) - valid_fields
    if unknown:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unknown columns found: {', '.join(sorted(unknown))}."
        )

    parsed = []
    for row_index, row in enumerate(rows[1:], start=2):
        if not any(cell is not None for cell in row):
            continue

        row_data = {}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            row_data[header] = row[col_index] if col_index < len(row) else None

        parsed.append((row_index, row_data))

    return parsed


# ==========================================
# CREATE - CREAR MATRÍCULA
# ==========================================
@router.post("/", response_model=EnrollmentResponse)
def create_enrollment(
    data: EnrollmentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Crear una nueva matrícula.
    Permisos:
    - Admin: puede crear cualquier matrícula
    - Profesor TUTOR: puede crear matrículas para estudiantes en las course offerings que tutoriza
    - Estudiante: solo puede crear su propia matrícula
    """
    
    # Verificar permisos según rol
    if current_user.role == "admin":
        # Admin puede todo
        pass
    
    elif current_user.role == "teacher":
        # Profesor: verificar si es tutor de alguna de las course offerings
        is_tutor = is_teacher_tutor_of_offerings(db, current_user.id, data.offering_ids)
        
        if not is_tutor:
            raise HTTPException(
                403, 
                "No tienes permiso para crear matrículas. Solo administradores o profesores tutores de la asignatura pueden hacerlo."
            )
        
        # El profesor tutor no puede matricularse a sí mismo
        if current_user.id == data.student_id:
            raise HTTPException(403, "Un profesor no puede matricularse a sí mismo como estudiante")
    
    elif current_user.id == data.student_id:
        # El propio estudiante puede matricularse
        pass
    
    else:
        raise HTTPException(403, "No tienes permiso para crear esta matrícula")
    
    # Verificar que el estudiante existe
    student = db.query(User).filter(User.id == data.student_id).first()
    if not student:
        raise HTTPException(404, f"Student with id {data.student_id} not found")
    
    # Verificar que el estudiante tiene rol 'student'
    if student.role != "student":
        raise HTTPException(400, f"User {data.student_id} is not a student")
    
    # Verificar que el año académico existe
    academic_year = db.query(AcademicYear).filter(AcademicYear.id == data.academic_year_id).first()
    if not academic_year:
        raise HTTPException(404, f"Academic year with id {data.academic_year_id} not found")
    
    # Verificar que todas las course offerings existen
    for offering_id in data.offering_ids:
        offering = db.query(CourseOffering).filter(CourseOffering.id == offering_id).first()
        if not offering:
            raise HTTPException(404, f"Offering {offering_id} not found")
        
        # Verificar que el año académico de la offering coincide
        if offering.academic_year_id != data.academic_year_id:
            raise HTTPException(400, f"Offering {offering_id} belongs to a different academic year")
    
    # Verificar que no existe una matrícula duplicada para el mismo estudiante y año
    existing_enrollment = db.query(Enrollment).filter(
        Enrollment.student_id == data.student_id,
        Enrollment.academic_year_id == data.academic_year_id
    ).first()
    
    if existing_enrollment:
        raise HTTPException(400, "El estudiante ya tiene una matrícula para este año académico")
    
    # Crear matrícula
    enrollment = Enrollment(
        student_id=data.student_id,
        academic_year_id=data.academic_year_id,
        enrollment_date=date.today()
    )

    db.add(enrollment)
    db.flush()

    # Insertar relaciones con course_offerings
    for offering_id in data.offering_ids:
        db.add(EnrollmentDetail(
            enrollment_id=enrollment.id,
            offering_id=offering_id
        ))

    db.commit()
    db.refresh(enrollment)
    
    return enrollment


# ==========================================
# GET ALL - LISTAR TODAS LAS MATRÍCULAS
# ==========================================
@router.get("/", response_model=list[EnrollmentResponse])
def get_all(
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_current_user)
):
    """
    Obtener todas las matrículas.
    Permisos:
    - Admin: todas las matrículas
    - Profesor TUTOR: solo matrículas de sus course offerings
    - Estudiante: solo sus propias matrículas
    """
    
    if current_user.role == "admin":
        enrollments = db.query(Enrollment).all()
    
    elif current_user.role == "teacher":
        # Obtener offerings donde el profesor es tutor
        tutor_offering_ids = get_tutor_offering_ids(db, current_user.id)
        
        if not tutor_offering_ids:
            return []
        
        # Obtener matrículas que incluyan esas offerings
        enrollments = db.query(Enrollment).join(
            EnrollmentDetail, Enrollment.id == EnrollmentDetail.enrollment_id
        ).filter(
            EnrollmentDetail.offering_id.in_(tutor_offering_ids)
        ).distinct().all()
    
    elif current_user.role == "student":
        enrollments = db.query(Enrollment).filter(
            Enrollment.student_id == current_user.id
        ).all()
    
    else:
        raise HTTPException(403, "Not authorized")
    
    return enrollments

# ==========================================
# DELETE - ELIMINAR MATRÍCULA
# ==========================================
@router.delete("/{enrollment_id}")
def delete_enrollment(
    enrollment_id: int, 
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_current_user)
):
    """
    Eliminar una matrícula.
    Permisos:
    - Admin: puede eliminar cualquier matrícula
    - Profesor TUTOR: puede eliminar matrículas de sus course offerings
    - Estudiante: puede eliminar su propia matrícula
    """
    
    enrollment = db.query(Enrollment).filter(Enrollment.id == enrollment_id).first()
    if not enrollment:
        raise HTTPException(404, "Enrollment not found")
    
    # Verificar permisos
    allowed = False
    
    if current_user.role == "admin":
        allowed = True
    elif current_user.id == enrollment.student_id:
        allowed = True
    elif current_user.role == "teacher":
        # Verificar si el profesor es tutor de alguna offering de esta matrícula
        details = db.query(EnrollmentDetail).filter(
            EnrollmentDetail.enrollment_id == enrollment_id
        ).all()
        offering_ids = [d.offering_id for d in details]
        allowed = is_teacher_tutor_of_offerings(db, current_user.id, offering_ids)
    
    if not allowed:
        raise HTTPException(403, "No tienes permiso para eliminar esta matrícula")

    db.query(EnrollmentDetail).filter(EnrollmentDetail.enrollment_id == enrollment_id).delete()
    db.delete(enrollment)
    db.commit()

    return {"message": "Enrollment deleted successfully"}


# ==========================================
# PATCH - ACTUALIZAR MATRÍCULA PARCIALMENTE
# ==========================================
@router.patch("/{enrollment_id}")
def patch_enrollment(
    enrollment_id: int,
    data: EnrollmentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Actualizar parcialmente una matrícula.
    Permisos:
    - Admin: puede modificar cualquier matrícula
    - Profesor TUTOR: puede modificar matrículas de sus course offerings
    """
    
    enrollment = db.query(Enrollment).filter(Enrollment.id == enrollment_id).first()

    if not enrollment:
        raise HTTPException(404, "Enrollment not found")
    
    # Verificar permisos
    allowed = False
    
    if current_user.role == "admin":
        allowed = True
    elif current_user.role == "teacher":
        # Verificar si el profesor es tutor de alguna offering de esta matrícula
        details = db.query(EnrollmentDetail).filter(
            EnrollmentDetail.enrollment_id == enrollment_id
        ).all()
        offering_ids = [d.offering_id for d in details]
        allowed = is_teacher_tutor_of_offerings(db, current_user.id, offering_ids)
    
    if not allowed:
        raise HTTPException(403, "No tienes permiso para modificar esta matrícula")

    for k, v in data.dict(exclude_unset=True).items():
        setattr(enrollment, k, v)

    db.commit()
    db.refresh(enrollment)
    return enrollment



# ==========================================
# GET CURRENT STUDENT ENROLLMENTS
# ==========================================
@router.get("/my-enrollments", response_model=list[EnrollmentWithCoursesResponse])
def get_my_enrollments(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Obtener las matrículas del estudiante actual (usando el token)"""
    if current_user.role not in ["student", "admin", "teacher"]:
        raise HTTPException(403, "Not allowed")
    
    # Obtener matrículas del usuario actual
    enrollments = db.query(Enrollment).filter(
        Enrollment.student_id == current_user.id
    ).all()
    
    result = []
    for enrollment in enrollments:
        academic_year_str = None
        if enrollment.academic_year:
            academic_year_str = f"{enrollment.academic_year.start_year}-{enrollment.academic_year.end_year}"
        else:
            academic_year_str = str(enrollment.academic_year_id)
        
        details = db.query(EnrollmentDetail).filter(
            EnrollmentDetail.enrollment_id == enrollment.id
        ).all()
        
        courses = []
        for detail in details:
            offering = db.query(CourseOffering).filter(
                CourseOffering.id == detail.offering_id
            ).first()
            
            if offering:
                subject = db.query(Subject).filter(Subject.id == offering.subject_id).first()
                if subject:
                    courses.append({
                        "offering_id": offering.id,
                        "course_id": subject.id,
                        "course_name": subject.name,
                        "course_description": subject.description or "",
                        "teachers": [],
                        "teacher_ids": [],
                        "academic_year": academic_year_str
                    })
        
        result.append({
            "id": enrollment.id,
            "student_id": enrollment.student_id,
            "academic_year_id": enrollment.academic_year_id,
            "academic_year": academic_year_str,
            "enrollment_date": enrollment.enrollment_date,
            "status": "active",
            "courses": courses
        })
    
    return result

# ==========================================
# GET BY STUDENT - MATRÍCULAS DE UN ESTUDIANTE
# ==========================================
@router.get("/student/{student_id}", response_model=list[EnrollmentWithCoursesResponse])
def get_enrollments_by_student(
    student_id: int, 
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_current_user)
):
    """Obtener todas las matrículas de un estudiante con detalles de cursos."""
    
    # Verificar permisos
    allowed = False
    
    if current_user.role == "admin":
        allowed = True
    elif current_user.id == student_id:
        allowed = True
    elif current_user.role == "teacher":
        # Verificar si el profesor es tutor de alguna offering donde el estudiante está matriculado
        enrollments = db.query(Enrollment).filter(Enrollment.student_id == student_id).all()
        for enrollment in enrollments:
            details = db.query(EnrollmentDetail).filter(
                EnrollmentDetail.enrollment_id == enrollment.id
            ).all()
            offering_ids = [d.offering_id for d in details]
            if is_teacher_tutor_of_offerings(db, current_user.id, offering_ids):
                allowed = True
                break
    
    if not allowed:
        raise HTTPException(403, "Not allowed - No tienes permiso para ver las matrículas de este estudiante")
    
    student = db.query(User).filter(User.id == student_id).first()
    if not student:
        raise HTTPException(404, f"Student with id {student_id} not found")
    
    enrollments = db.query(Enrollment).filter(Enrollment.student_id == student_id).all()
    
    result = []
    for enrollment in enrollments:
        academic_year_str = None
        if enrollment.academic_year:
            academic_year_str = f"{enrollment.academic_year.start_year}-{enrollment.academic_year.end_year}"
        else:
            academic_year_str = str(enrollment.academic_year_id)
        
        details = db.query(EnrollmentDetail).filter(
            EnrollmentDetail.enrollment_id == enrollment.id
        ).all()
        
        courses = []
        for detail in details:
            offering = db.query(CourseOffering).filter(
                CourseOffering.id == detail.offering_id
            ).first()
            
            if offering:
                subject = db.query(Subject).filter(Subject.id == offering.subject_id).first()
                if subject:
                    teachers = db.query(TeacherAssignment).filter(
                        TeacherAssignment.course_offering_id == offering.id
                    ).all()
                    
                    teacher_names = []
                    teacher_ids = []
                    for ta in teachers:
                        teacher = db.query(User).filter(User.id == ta.professor_id).first()
                        if teacher:
                            teacher_names.append(teacher.name)
                            teacher_ids.append(teacher.id)
                    
                    courses.append({
                        "offering_id": offering.id,
                        "course_id": subject.id,
                        "course_name": subject.name,
                        "course_description": subject.description or "",
                        "teachers": teacher_names,
                        "teacher_ids": teacher_ids,
                        "academic_year": academic_year_str
                    })
        
        result.append({
            "id": enrollment.id,
            "student_id": enrollment.student_id,
            "academic_year_id": enrollment.academic_year_id,
            "academic_year": academic_year_str,
            "enrollment_date": enrollment.enrollment_date,
            "status": "active",
            "courses": courses
        })
    
    return result


# ==========================================
# GET ENROLLMENTS BY COURSE OFFERING
# ==========================================
@router.get("/by-offering/{offering_id}")
def get_enrollments_by_offering(
    offering_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener todos los estudiantes matriculados en una course offering específica.
    Permisos:
    - Admin: puede ver cualquier offering
    - Profesor: puede ver offerings donde imparte la asignatura
    - Estudiante: puede ver compañeros matriculados en la misma offering
    """
    
    # Verificar que la offering existe
    offering = db.query(CourseOffering).filter(CourseOffering.id == offering_id).first()
    if not offering:
        raise HTTPException(404, "Course offering not found")
    
    # Verificar permisos según el rol
    if current_user.role == "admin":
        # Admin puede ver todo
        pass
    
    elif current_user.role == "teacher":
        # Verificar si el profesor está asignado a esta course offering
        teacher_assignment = db.query(TeacherAssignment).filter(
            TeacherAssignment.professor_id == current_user.id,
            TeacherAssignment.course_offering_id == offering_id
        ).first()
        
        if not teacher_assignment:
            raise HTTPException(403, "No tienes permiso para ver los estudiantes de esta oferta")
    
    elif current_user.role == "student":
        # Verificar si el estudiante está matriculado en esta offering
        is_enrolled = db.query(EnrollmentDetail).join(
            Enrollment, EnrollmentDetail.enrollment_id == Enrollment.id
        ).filter(
            Enrollment.student_id == current_user.id,
            EnrollmentDetail.offering_id == offering_id
        ).first()
        
        if not is_enrolled:
            raise HTTPException(403, "No estás matriculado en esta asignatura, no puedes ver los estudiantes")
    
    else:
        raise HTTPException(403, "No tienes permiso para ver los estudiantes de esta oferta")
    
    # Obtener estudiantes
    students = db.query(User).join(
        Enrollment, User.id == Enrollment.student_id
    ).join(
        EnrollmentDetail, Enrollment.id == EnrollmentDetail.enrollment_id
    ).filter(
        EnrollmentDetail.offering_id == offering_id,
        User.role == "student"
    ).distinct().all()
    
    result = []
    for student in students:
        # Obtener la matrícula específica para esta offering
        enrollment = db.query(Enrollment).join(
            EnrollmentDetail, Enrollment.id == EnrollmentDetail.enrollment_id
        ).filter(
            Enrollment.student_id == student.id,
            EnrollmentDetail.offering_id == offering_id
        ).first()
        
        result.append({
            "id": student.id,
            "name": student.name,
            "email": student.email,
            "enrollment_id": enrollment.id if enrollment else None,
            "enrollment_date": enrollment.enrollment_date if enrollment else None
        })
    
    return {
        "offering_id": offering_id,
        "subject_name": offering.subject.name if offering.subject else None,
        "academic_year": f"{offering.academic_year.start_year}-{offering.academic_year.end_year}" if offering.academic_year else None,
        "students_count": len(result),
        "students": result
    }

# ==========================================
# TEMPLATE
# ==========================================
@router.get("/template", status_code=200)
def download_enrollments_template(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Download an Excel template for bulk enrollment upload.
    Columns: academic_year_start, academic_year_end, name, email, role, offering_subject, 
             offering_academic_year_start, offering_academic_year_end, password.
    
    Permisos:
    - Admin: puede descargar
    - Profesor TUTOR: puede descargar (para preparar uploads)
    """
    
    if current_user.role not in ["admin", "teacher"]:
        raise HTTPException(403, "Solo administradores o profesores pueden descargar la plantilla")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "plantilla_matricula"
    headers = [
        "academic_year_start",
        "academic_year_end",
        "name",
        "email",
        "role",
        "offering_subject",
        "offering_academic_year_start",
        "offering_academic_year_end",
        "password"
    ]
    ws.append(headers)
    ws.append([2025, 2026, "Juan Pérez", "juan@example.com", "student", "Algorithms", 2025, 2026, ""])
    ws.append([2025, 2026, "Ana López", "", "student", "Data Structures", 2025, 2026, ""])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = "plantilla_excel_matricula_final.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==========================================
# UPLOAD ENROLLMENTS (Excel) - Admin y Profesores Tutores
# ==========================================

# ==========================================
# UPLOAD ENROLLMENTS (Excel) - Admin y Profesores Tutores
# ==========================================

@router.post("/upload", response_model=BulkEnrollmentResult, status_code=status.HTTP_201_CREATED)
def upload_enrollments(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Upload an Excel file (.xlsx) using the same format as the template.
    Expected columns:
        - academic_year_start, academic_year_end (required)
        - email (required) or name (optional, with email)
        - offering_subject (required) + offering_academic_year_start, offering_academic_year_end
    Optional: name, role, password (for user creation)
    
    Permisos:
    - Admin: puede hacer uploads a cualquier course offering
    - Profesor TUTOR: solo puede hacer uploads a las course offerings donde es tutor
    """
    
    # Verificar permisos básicos
    if current_user.role not in ["admin", "teacher"]:
        raise HTTPException(403, "Solo administradores o profesores pueden realizar uploads de matrículas")
    
    # Si es profesor, obtener las offerings donde es tutor
    tutor_offering_ids = []
    if current_user.role == "teacher":
        tutor_offering_ids = get_tutor_offering_ids(db, current_user.id)
        if not tutor_offering_ids:
            raise HTTPException(403, "No eres tutor de ninguna asignatura. No puedes realizar uploads.")
    
    # Parsear el Excel usando la función existente (ya soporta todas las columnas de la plantilla)
    rows = _parse_excel_enrollments_with_users(file)
    created_enrollments = []
    errors = []

    for row_number, row_data in rows:
        # Normalizar strings
        row_data = {k: (str(v).strip() if v is not None else None) for k, v in row_data.items()}

        # --- Resolver academic_year_id desde start/end ---
        academic_year_id = None
        if row_data.get("academic_year_start") and row_data.get("academic_year_end"):
            try:
                start_year = int(float(row_data.get("academic_year_start")))
                end_year = int(float(row_data.get("academic_year_end")))
                ay = db.query(AcademicYear).filter(
                    AcademicYear.start_year == start_year,
                    AcademicYear.end_year == end_year
                ).first()
                if not ay:
                    errors.append({"row": row_number, "status": "error", "detail": f"Año académico {start_year}-{end_year} no encontrado"})
                    continue
                academic_year_id = ay.id
            except (ValueError, TypeError):
                errors.append({"row": row_number, "status": "error", "detail": "academic_year_start o academic_year_end inválidos"})
                continue
        else:
            errors.append({"row": row_number, "status": "error", "detail": "Faltan academic_year_start y academic_year_end"})
            continue

        # --- Resolver estudiante (por email o nombre) ---
        email = row_data.get("email")
        name = row_data.get("name")
        role = row_data.get("role", "student")
        password = row_data.get("password")

        user_obj = None
        if email:
            email = email.lower()
            user_obj = db.query(User).filter(User.email == email).first()
            if not user_obj:
                # Solo admin puede crear nuevos usuarios
                if current_user.role != "admin":
                    errors.append({"row": row_number, "status": "error", "detail": f"Usuario con email {email} no existe. Solo administradores pueden crear usuarios durante el upload."})
                    continue
                # Crear usuario
                temp_password = password or generate_random_password(12)
                new_user = User(
                    email=email,
                    name=name or "",
                    hashed_password=hash_password(temp_password),
                    role=role,
                    must_change_password=True,
                )
                try:
                    db.add(new_user)
                    db.flush()
                    send_temporary_password(new_user.email, temp_password, new_user.name)
                    db.commit()
                    db.refresh(new_user)
                    user_obj = new_user
                except Exception as exc:
                    db.rollback()
                    errors.append({"row": row_number, "status": "error", "detail": f"No se pudo crear el usuario: {str(exc)}"})
                    continue
        elif name:
            # Buscar por nombre (solo si es admin)
            if current_user.role != "admin":
                errors.append({"row": row_number, "status": "error", "detail": "Solo administradores pueden usar el campo 'name' sin email"})
                continue
            matches = db.query(User).filter(User.name == name).all()
            if len(matches) == 1:
                user_obj = matches[0]
            elif len(matches) > 1:
                errors.append({"row": row_number, "status": "error", "detail": f"Múltiples usuarios con nombre '{name}'. Usa el email para desambiguar."})
                continue
            else:
                # Crear usuario con email temporal
                domain = os.getenv("PLACEHOLDER_EMAIL_DOMAIN", "noemail.local")
                slug = re.sub(r"[^a-z0-9]+", ".", name.lower()).strip(".")
                unique_suffix = f"{int(time.time())}{secrets.token_hex(2)}"
                placeholder_email = f"{slug}.{unique_suffix}@{domain}"
                temp_password = password or generate_random_password(12)
                new_user = User(
                    email=placeholder_email,
                    name=name,
                    hashed_password=hash_password(temp_password),
                    role=role,
                    must_change_password=True,
                )
                try:
                    db.add(new_user)
                    db.flush()
                    send_temporary_password(new_user.email, temp_password, new_user.name)
                    db.commit()
                    db.refresh(new_user)
                    user_obj = new_user
                except Exception as exc:
                    db.rollback()
                    errors.append({"row": row_number, "status": "error", "detail": f"No se pudo crear el usuario: {str(exc)}"})
                    continue
        else:
            errors.append({"row": row_number, "status": "error", "detail": "No se proporcionó email ni nombre"})
            continue

        # Verificar que sea estudiante
        if user_obj.role != "student":
            errors.append({"row": row_number, "status": "error", "detail": f"El usuario {user_obj.email} no tiene rol 'student'"})
            continue

        # --- Resolver offering_ids desde offering_subject + offering_academic_year_start/end ---
        offering_subject = row_data.get("offering_subject")
        offering_ay_start = row_data.get("offering_academic_year_start")
        offering_ay_end = row_data.get("offering_academic_year_end")

        if not offering_subject or not offering_ay_start or not offering_ay_end:
            errors.append({"row": row_number, "status": "error", "detail": "Faltan offering_subject, offering_academic_year_start o offering_academic_year_end"})
            continue

        try:
            offering_start = int(float(offering_ay_start))
            offering_end = int(float(offering_ay_end))
            subject = db.query(Subject).filter(Subject.name == offering_subject).first()
            if not subject:
                errors.append({"row": row_number, "status": "error", "detail": f"Asignatura '{offering_subject}' no encontrada"})
                continue
            ay_offering = db.query(AcademicYear).filter(
                AcademicYear.start_year == offering_start,
                AcademicYear.end_year == offering_end
            ).first()
            if not ay_offering:
                errors.append({"row": row_number, "status": "error", "detail": f"Año académico {offering_start}-{offering_end} no encontrado para la oferta"})
                continue
            offerings = db.query(CourseOffering).filter(
                CourseOffering.subject_id == subject.id,
                CourseOffering.academic_year_id == ay_offering.id
            ).all()
            if not offerings:
                errors.append({"row": row_number, "status": "error", "detail": f"No hay ofertas para '{offering_subject}' en {offering_start}-{offering_end}"})
                continue
            offering_ids = [o.id for o in offerings]
        except (ValueError, TypeError):
            errors.append({"row": row_number, "status": "error", "detail": "Valores inválidos en offering_academic_year_start o offering_academic_year_end"})
            continue

        # Si es profesor, verificar que todas las offering_ids están en sus tutorías
        if current_user.role == "teacher":
            invalid = [oid for oid in offering_ids if oid not in tutor_offering_ids]
            if invalid:
                errors.append({"row": row_number, "status": "error", "detail": f"No tienes permiso para matricular en las offerings: {invalid} (no eres tutor)"})
                continue

        # Verificar si ya existe matrícula para este estudiante en el mismo año académico
        existing = db.query(Enrollment).filter(
            Enrollment.student_id == user_obj.id,
            Enrollment.academic_year_id == academic_year_id
        ).first()
        if existing:
            errors.append({"row": row_number, "status": "error", "detail": f"El estudiante {user_obj.email} ya está matriculado en el año académico {academic_year_id}"})
            continue

        # Crear matrícula y detalles
        try:
            enrollment = Enrollment(
                student_id=user_obj.id,
                academic_year_id=academic_year_id,
                enrollment_date=date.today()
            )
            db.add(enrollment)
            db.flush()

            for oid in offering_ids:
                db.add(EnrollmentDetail(enrollment_id=enrollment.id, offering_id=oid))

            db.commit()
            db.refresh(enrollment)
            created_enrollments.append(enrollment)
        except Exception as exc:
            db.rollback()
            errors.append({"row": row_number, "status": "error", "detail": f"Error al crear matrícula: {str(exc)}"})

    return BulkEnrollmentResult(
        created_count=len(created_enrollments),
        created=created_enrollments,
        errors=errors
    )


# ==========================================
# UPLOAD ENROLLMENTS WITH USERS (Excel) - Admin y Profesores Tutores
# ==========================================

@router.post("/upload-with-users", response_model=BulkEnrollmentResult, status_code=status.HTTP_201_CREATED)
def upload_enrollments_with_users(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Upload an Excel file (.xlsx) containing student info and enrollment details.
    Can use numeric IDs or human-readable identifiers. Supports two enrollment modes:
    1. By numeric IDs: academic_year_id, offering_ids
    2. By names: academic_year_start/academic_year_end, offering_subject + offering_academic_year_start/offering_academic_year_end
    
    Required: academic_year_id OR (academic_year_start and academic_year_end)
    Required: student_id OR email OR name
    Optional: role, password, offering_ids OR (offering_subject + offering_academic_year_start + offering_academic_year_end)
    
    Permisos:
    - Admin: puede hacer uploads a cualquier course offering
    - Profesor TUTOR: solo puede hacer uploads a las course offerings donde es tutor
    """
    
    # Verificar permisos básicos
    if current_user.role not in ["admin", "teacher"]:
        raise HTTPException(403, "Solo administradores o profesores pueden realizar uploads de matrículas")
    
    # Si es profesor, obtener las offerings donde es tutor
    tutor_offering_ids = []
    if current_user.role == "teacher":
        tutor_offering_ids = get_tutor_offering_ids(db, current_user.id)
        if not tutor_offering_ids:
            raise HTTPException(403, "No eres tutor de ninguna asignatura. No puedes realizar uploads.")
    
    rows = _parse_excel_enrollments_with_users(file)
    created_enrollments = []
    errors = []

    for row_number, row_data in rows:
        # normalize strings
        row_data = {k: (str(v).strip() if v is not None else None) for k, v in row_data.items()}

        # Resolve academic_year_id from either numeric id or start/end years
        academic_year_id = None
        if row_data.get("academic_year_id"):
            try:
                academic_year_id = int(float(row_data.get("academic_year_id")))
            except (ValueError, TypeError):
                errors.append({"row": row_number, "status": "error", "detail": "Invalid academic_year_id"})
                continue
        elif row_data.get("academic_year_start") and row_data.get("academic_year_end"):
            try:
                start_year = int(float(row_data.get("academic_year_start")))
                end_year = int(float(row_data.get("academic_year_end")))
                ay = db.query(AcademicYear).filter(
                    AcademicYear.start_year == start_year,
                    AcademicYear.end_year == end_year
                ).first()
                if not ay:
                    errors.append({"row": row_number, "status": "error", "detail": f"Academic year {start_year}-{end_year} not found"})
                    continue
                academic_year_id = ay.id
            except (ValueError, TypeError):
                errors.append({"row": row_number, "status": "error", "detail": "Invalid academic_year_start or academic_year_end"})
                continue
        else:
            errors.append({"row": row_number, "status": "error", "detail": "No valid academic year identifier provided"})
            continue

        student_id = None
        if row_data.get("student_id"):
            try:
                student_id = int(float(row_data.get("student_id")))
            except (ValueError, TypeError):
                errors.append({"row": row_number, "status": "error", "detail": "Invalid student_id"})
                continue

        # Determine the user object to use
        user_obj = None

        if student_id:
            user_obj = db.query(User).filter(User.id == student_id).first()
            if not user_obj:
                errors.append({"row": row_number, "status": "error", "detail": f"Student with id {student_id} not found"})
                continue
        else:
            email = row_data.get("email")
            name = row_data.get("name")

            if email:
                email = email.lower()
                user_obj = db.query(User).filter(User.email == email).first()
                if not user_obj:
                    # Solo admin puede crear nuevos usuarios
                    if current_user.role != "admin":
                        errors.append({
                            "row": row_number, 
                            "status": "error", 
                            "detail": "Solo administradores pueden crear nuevos usuarios durante el upload"
                        })
                        continue
                    
                    temp_password = row_data.get("password") or generate_random_password(12)
                    new_user = User(
                        email=email,
                        name=name or "",
                        password=hash_password(temp_password),
                        role=row_data.get("role") or "student",
                        needs_password_change=True,
                    )
                    try:
                        db.add(new_user)
                        db.flush()
                        send_temporary_password(new_user.email, temp_password, new_user.name)
                        db.commit()
                        db.refresh(new_user)
                        user_obj = new_user
                    except Exception as exc:
                        db.rollback()
                        errors.append({"row": row_number, "status": "error", "detail": f"Unable to create user: {str(exc)}"})
                        continue
            elif name:
                # Solo admin puede crear usuarios por nombre
                if current_user.role != "admin":
                    errors.append({
                        "row": row_number,
                        "status": "error",
                        "detail": "Solo administradores pueden crear usuarios por nombre"
                    })
                    continue
                    
                matches = db.query(User).filter(User.name == name).all()
                if len(matches) == 1:
                    user_obj = matches[0]
                elif len(matches) > 1:
                    errors.append({
                        "row": row_number,
                        "status": "error",
                        "detail": "Multiple users found with this name. Provide the email to disambiguate."
                    })
                    continue
                else:
                    domain = os.getenv("PLACEHOLDER_EMAIL_DOMAIN", "noemail.local")
                    slug = re.sub(r"[^a-z0-9]+", ".", name.lower()).strip(".")
                    unique_suffix = f"{int(time.time())}{secrets.token_hex(2)}"
                    placeholder_email = f"{slug}.{unique_suffix}@{domain}"

                    temp_password = row_data.get("password") or generate_random_password(12)
                    new_user = User(
                        email=placeholder_email,
                        name=name,
                        password=hash_password(temp_password),
                        role=row_data.get("role") or "student",
                        needs_password_change=True,
                    )
                    try:
                        db.add(new_user)
                        db.flush()
                        send_temporary_password(new_user.email, temp_password, new_user.name)
                        db.commit()
                        db.refresh(new_user)
                        user_obj = new_user
                    except Exception as exc:
                        db.rollback()
                        errors.append({"row": row_number, "status": "error", "detail": f"Unable to create user: {str(exc)}"})
                        continue
            else:
                errors.append({"row": row_number, "status": "error", "detail": "No student identifier provided"})
                continue

        # Verificar que el usuario es estudiante
        if user_obj.role != "student":
            errors.append({"row": row_number, "status": "error", "detail": f"User {user_obj.id} is not a student"})
            continue

        # verify academic year
        academic_year = db.query(AcademicYear).filter(AcademicYear.id == academic_year_id).first()
        if not academic_year:
            errors.append({"row": row_number, "status": "error", "detail": f"Academic year {academic_year_id} not found"})
            continue

        # check existing enrollment
        existing = db.query(Enrollment).filter(
            Enrollment.student_id == user_obj.id,
            Enrollment.academic_year_id == academic_year_id
        ).first()
        if existing:
            errors.append({"row": row_number, "status": "error", "detail": "Enrollment already exists"})
            continue

        # Resolve offering IDs
        offering_ids = []
        offering_ids_str = row_data.get("offering_ids") or ""
        offering_subject = row_data.get("offering_subject") or ""
        offering_ay_start = row_data.get("offering_academic_year_start")
        offering_ay_end = row_data.get("offering_academic_year_end")

        if offering_ids_str:
            try:
                offering_ids = [int(float(x.strip())) for x in offering_ids_str.split(",") if x.strip()]
            except Exception:
                errors.append({"row": row_number, "status": "error", "detail": "Invalid offering_ids format"})
                continue
        elif offering_subject and offering_ay_start and offering_ay_end:
            try:
                offering_start = int(float(offering_ay_start))
                offering_end = int(float(offering_ay_end))
                subject = db.query(Subject).filter(Subject.name == offering_subject).first()
                if not subject:
                    errors.append({"row": row_number, "status": "error", "detail": f"Subject '{offering_subject}' not found"})
                    continue
                ay_for_offering = db.query(AcademicYear).filter(
                    AcademicYear.start_year == offering_start,
                    AcademicYear.end_year == offering_end
                ).first()
                if not ay_for_offering:
                    errors.append({"row": row_number, "status": "error", "detail": f"Academic year {offering_start}-{offering_end} not found for offering"})
                    continue
                offerings = db.query(CourseOffering).filter(
                    CourseOffering.subject_id == subject.id,
                    CourseOffering.academic_year_id == ay_for_offering.id
                ).all()
                if not offerings:
                    errors.append({"row": row_number, "status": "error", "detail": f"No offerings found for subject '{offering_subject}' in {offering_start}-{offering_end}"})
                    continue
                offering_ids = [o.id for o in offerings]
            except (ValueError, TypeError):
                errors.append({"row": row_number, "status": "error", "detail": "Invalid offering_academic_year_start or offering_academic_year_end"})
                continue

        # Si es profesor, verificar que todas las offering_ids están en sus tutorías
        if current_user.role == "teacher":
            invalid_offerings = [oid for oid in offering_ids if oid not in tutor_offering_ids]
            if invalid_offerings:
                errors.append({
                    "row": row_number,
                    "status": "error",
                    "detail": f"No tienes permiso para matricular en las offerings: {invalid_offerings} (no eres tutor)"
                })
                continue

        # create enrollment and details
        try:
            enrollment = Enrollment(
                student_id=user_obj.id,
                academic_year_id=academic_year_id,
                enrollment_date=date.today()
            )
            db.add(enrollment)
            db.flush()

            for offering_id in offering_ids:
                offering = db.query(CourseOffering).filter(CourseOffering.id == offering_id).first()
                if not offering:
                    errors.append({"row": row_number, "status": "warning", "detail": f"Offering {offering_id} not found, skipped"})
                    continue
                db.add(EnrollmentDetail(enrollment_id=enrollment.id, offering_id=offering_id))

            db.commit()
            db.refresh(enrollment)
            created_enrollments.append(enrollment)
        except Exception as exc:
            db.rollback()
            errors.append({"row": row_number, "status": "error", "detail": str(exc)})

    return BulkEnrollmentResult(created_count=len(created_enrollments), created=created_enrollments, errors=errors)


# ==========================================
# SEARCH STUDENTS BY NAME (para profesores)
# ==========================================
@router.get("/search-students")
def search_students(
    q: str,
    offering_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Buscar estudiantes por nombre.
    Permisos:
    - Admin: puede buscar cualquier estudiante
    - Profesor: solo estudiantes matriculados en sus course offerings
    """
    
    # Query base
    query = db.query(User).filter(User.role == "student")
    
    # Si es profesor, aplicar filtros según el caso
    if current_user.role == "teacher":
        # Determinar qué offering_ids usar
        target_offering_ids = []
        
        if offering_id:
            # Si se especifica offering_id, verificar permiso y usarlo directamente
            teacher_assignment = db.query(TeacherAssignment).filter(
                TeacherAssignment.professor_id == current_user.id,
                TeacherAssignment.course_offering_id == offering_id
            ).first()
            
            if not teacher_assignment:
                raise HTTPException(403, "No tienes permiso para ver esta oferta")
            
            target_offering_ids = [offering_id]
        else:
            # Obtener todas las offerings del profesor
            teacher_offerings = db.query(TeacherAssignment.course_offering_id).filter(
                TeacherAssignment.professor_id == current_user.id
            ).distinct().all()
            target_offering_ids = [o[0] for o in teacher_offerings]
            
            if not target_offering_ids:
                return {"students": [], "total": 0}
        
        # Aplicar el filtro de matriculación (solo una vez)
        query = query.join(
            Enrollment, User.id == Enrollment.student_id
        ).join(
            EnrollmentDetail, Enrollment.id == EnrollmentDetail.enrollment_id
        ).filter(
            EnrollmentDetail.offering_id.in_(target_offering_ids)
        ).distinct()
    
    elif current_user.role != "admin":
        # Ni admin ni teacher -> no autorizado
        raise HTTPException(403, "Not authorized")
    
    # Filtrar por término de búsqueda
    if q and q.strip():
        if len(q.strip()) == 1:
            query = query.filter(
                or_(
                    User.name.ilike(f"{q.strip()}%"),
                    User.email.ilike(f"{q.strip()}%")
                )
            )
        else:
            query = query.filter(
                or_(
                    User.name.ilike(f"%{q}%"),
                    User.email.ilike(f"%{q}%")
                )
            )
    
    students = query.limit(50).all()
    
    result = []
    for student in students:
        result.append({
            "id": student.id,
            "name": student.name,
            "email": student.email
        })
    
    return {
        "students": result,
        "total": len(result)
    }
# ==========================================
# SEARCH STUDENTS NOT ENROLLED IN COURSE OFFERING
# ==========================================
@router.get("/search-not-enrolled")
def search_students_not_enrolled(
    q: str,
    offering_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Buscar estudiantes NO matriculados en una course offering específica."""
    
    # Verificar permisos
    if current_user.role != "admin":
        teacher_assignment = db.query(TeacherAssignment).filter(
            TeacherAssignment.professor_id == current_user.id,
            TeacherAssignment.course_offering_id == offering_id
        ).first()
        if not teacher_assignment:
            raise HTTPException(403, "No tienes permiso para ver esta oferta")
    
    # Verificar que la offering existe
    offering = db.query(CourseOffering).filter(CourseOffering.id == offering_id).first()
    if not offering:
        raise HTTPException(404, "Course offering not found")
    
    # MÉTODO SIMPLE: Obtener IDs de estudiantes YA matriculados
    # Paso 1: Obtener enrollment_ids de esta offering
    enrollment_details = db.query(EnrollmentDetail.enrollment_id).filter(
        EnrollmentDetail.offering_id == offering_id
    ).all()
    
    enrollment_ids = [ed[0] for ed in enrollment_details]
    
    # Paso 2: Obtener student_ids de esos enrollments
    enrolled_student_ids = []
    if enrollment_ids:
        enrollments = db.query(Enrollment.student_id).filter(
            Enrollment.id.in_(enrollment_ids)
        ).all()
        enrolled_student_ids = [e[0] for e in enrollments]
    
    # Paso 3: Buscar estudiantes NO matriculados
    query = db.query(User).filter(User.role == "student")
    
    # Excluir los que ya están matriculados
    if enrolled_student_ids:
        query = query.filter(User.id.notin_(enrolled_student_ids))
    
    # Filtrar por término de búsqueda
    if q and q.strip():
        if len(q.strip()) == 1:
            query = query.filter(
                or_(  # ← Usar or_ directamente, no db.or_
                    User.name.ilike(f"{q.strip()}%"),
                    User.email.ilike(f"{q.strip()}%")
                )
            )
        else:
            query = query.filter(
                or_(  # ← Usar or_ directamente, no db.or_
                    User.name.ilike(f"%{q}%"),
                    User.email.ilike(f"%{q}%")
                )
            )
    
    students = query.limit(50).all()
    
    # Obtener años académicos
    academic_years = db.query(AcademicYear).all()
    
    result = []
    for student in students:
        result.append({
            "student_id": student.id,
            "name": student.name,
            "email": student.email
        })
    
    return {
        "offering_id": offering_id,
        "students": result,
        "total": len(result),
        "academic_years": [
            {"id": ay.id, "name": f"{ay.start_year}-{ay.end_year}"}
            for ay in academic_years
        ]
    }
# ==========================================
# GET BY ID - OBTENER MATRÍCULA POR ID
# ==========================================
@router.get("/{enrollment_id}", response_model=EnrollmentWithCoursesResponse)
def get_enrollment_by_id(
    enrollment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Obtener una matrícula específica por ID"""
    enrollment = db.query(Enrollment).filter(Enrollment.id == enrollment_id).first()
    
    if not enrollment:
        raise HTTPException(404, "Enrollment not found")
    
    # Verificar permisos
    allowed = False
    
    if current_user.role == "admin":
        allowed = True
    elif current_user.role == "teacher":
        # Verificar si el profesor es tutor de alguna offering de esta matrícula
        details = db.query(EnrollmentDetail).filter(
            EnrollmentDetail.enrollment_id == enrollment_id
        ).all()
        offering_ids = [d.offering_id for d in details]
        allowed = is_teacher_tutor_of_offerings(db, current_user.id, offering_ids)
    elif current_user.id == enrollment.student_id:
        allowed = True
    
    if not allowed:
        raise HTTPException(403, "Not allowed")
    
    # Obtener año académico como string
    academic_year_str = None
    if enrollment.academic_year:
        academic_year_str = f"{enrollment.academic_year.start_year}-{enrollment.academic_year.end_year}"
    else:
        academic_year_str = str(enrollment.academic_year_id)
    
    # Obtener detalles
    details = db.query(EnrollmentDetail).filter(
        EnrollmentDetail.enrollment_id == enrollment.id
    ).all()
    
    courses = []
    for detail in details:
        offering = db.query(CourseOffering).filter(
            CourseOffering.id == detail.offering_id
        ).first()
        
        if offering:
            subject = db.query(Subject).filter(Subject.id == offering.subject_id).first()
            if subject:
                teachers = db.query(TeacherAssignment).filter(
                    TeacherAssignment.course_offering_id == offering.id
                ).all()
                
                teacher_names = []
                teacher_ids = []
                for ta in teachers:
                    teacher = db.query(User).filter(User.id == ta.professor_id).first()
                    if teacher:
                        teacher_names.append(teacher.name)
                        teacher_ids.append(teacher.id)
                
                courses.append({
                    "offering_id": offering.id,
                    "course_id": subject.id,
                    "course_name": subject.name,
                    "course_description": subject.description or "",
                    "teachers": teacher_names,
                    "teacher_ids": teacher_ids,
                    "academic_year": academic_year_str
                })
    
    return {
        "id": enrollment.id,
        "student_id": enrollment.student_id,
        "academic_year_id": enrollment.academic_year_id,
        "academic_year": academic_year_str,
        "enrollment_date": enrollment.enrollment_date,
        "status": "active",
        "courses": courses
    }

